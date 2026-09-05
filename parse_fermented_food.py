# -*- coding: utf-8 -*-
"""解析《发酵食品挥发性物质文献数据库.xlsx》为 fermented_food_db.json。
- 主表(化合物主表)：中文名/英文名/CAS/分类/风味描述汇总
- 阈值记录(阈值记录)：CAS/英文名 + 阈值原值(去脚注a/b) + 单位 + 介质 + 风味描述
同 CAS / 同归一化英文名 视为同物，合并同义名、取最优中文名。
"""
import openpyxl, re, json, os, importlib.util

HERE = os.path.dirname(os.path.abspath(__file__))
spec = importlib.util.spec_from_file_location("fc", os.path.join(HERE, "flavor_core.py"))
fc = importlib.util.module_from_spec(spec); spec.loader.exec_module(fc)

SRC = "/root/uploads/1787627335825691829-发酵食品挥发性物质文献数据库.xlsx"
OUT = os.path.join(HERE, "fermented_food_db.json")

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

def main_en(en):
    en = (en or "").strip()
    return re.sub(r"\s*\(.*?\)\s*", " ", en).strip()

def extract_val(s):
    if s is None:
        return None
    s = str(s).strip().lower()
    s = re.sub(r"\s*[ab]$", "", s)        # 去脚注 a/b
    s = s.replace(",", "")
    m = re.search(r"[-+]?[\d.]+", s)
    return float(m.group()) if m else None

def unit_to_thr(u):
    u = (u or "").strip()
    if u in ("μg/L", "ug/L", "µg/L"):
        return "≈%.0f" % float("1") + " μg/L" if False else None  # 占位
    return u

def med_from(u, medium):
    u = (u or "").strip()
    if u in ("μg/L", "ug/L", "µg/L"):
        return "水(ppb)"
    if u in ("mg/L", "mg/kg"):
        return "水(ppm)"
    if u in ("ng/L",):
        return "水(ppt)"
    return ((medium or "")[:18].strip() or u or "—")

# ---- 阈值记录：按 CAS / 归一化英文名 取首个有效数值 ----
wr = wb["阈值记录"]
trows = list(wr.iter_rows(values_only=True))[1:]
thr_map = {}
for r in trows:
    if not r:
        continue
    en = (r[3] or "").strip(); cas = (r[4] or "").strip()
    cas = "" if cas in ("", "—", None) else cas.strip()
    if not en and not cas:
        continue
    val = extract_val(r[6]); u = r[7]; medium = r[8]; fla = (r[9] or "").strip()
    key = ("CAS:" + cas) if cas else ("EN:" + fc.normalize(main_en(en)))
    if key not in thr_map:
        thr_map[key] = dict(
            thr_val=val,
            thr=("≈%.4g" % val) if val is not None else None,
            med=med_from(u, medium),
            fla=fla,
        )

# ---- 主表：聚合同物异名 ----
ws = wb["化合物主表"]
mrows = list(ws.iter_rows(values_only=True))[1:]
agg = {}
for r in mrows:
    if not r:
        continue
    cn = (r[1] or "").strip(); en = (r[2] or "").strip(); cas = (r[3] or "").strip()
    cat = (r[4] or "").strip(); fla = (r[6] or "").strip()
    if not en and not cas:
        continue
    cas = "" if cas in ("", "—", None) else cas.strip()
    men = main_en(en)
    is_pending = cn.startswith("（待核") or cn.startswith("(待核") or cn == "（待核中文名）"
    key = ("CAS:" + cas) if cas else ("EN:" + fc.normalize(men))
    g = agg.get(key)
    if not g:
        g = dict(en=men, raw_ens=[en], cas=cas, cn=cn, cat=cat, fla=fla,
                 pending=is_pending, seen_cn=set())
        agg[key] = g
    else:
        g["raw_ens"].append(en)
        g["fla"] = (g["fla"] + "；" + fla).strip("；") if fla and fla not in g["fla"] else g["fla"]
        # 中文名：优先取非待核的
        if is_pending and not g["pending"]:
            pass
        elif (not is_pending) and g["pending"]:
            g["cn"] = cn; g["pending"] = False; g["cat"] = cat or g["cat"]
        elif (not is_pending) and (not g["pending"]):
            if cn and len(cn) > len(g["cn"] or ""):
                g["cn"] = cn
        g["cat"] = cat or g["cat"]
    if cn and not is_pending:
        g["seen_cn"].add(cn)

def map_cat(raw):
    r = raw or ""
    if "醛" in r: return "醛类"
    if "酮" in r: return "酮类"
    if "酯" in r: return "酯类"
    if "酸" in r: return "酸类"
    if "醇" in r: return "醇类"
    if "萜" in r: return "萜烯类"
    return "其他"

def fmt_thr(val):
    if val is None:
        return None
    if val >= 100:
        return "≈%.0f" % val
    if val >= 1:
        return "≈%.2g" % val
    return "≈%.4g" % val

def clean_cn(cn):
    """仅保留中文字符，去除英文/数字/标点碎片。如 '2-Methoxy-苯酚' -> '苯酚'。"""
    if not cn:
        return ""
    zh = re.findall(r"[一-鿿]", cn)
    return "".join(zh)

items = []
for key, g in agg.items():
    t = thr_map.get(key, {})
    cn = g["cn"] if not g["pending"] else ""
    # 风味描述：主表汇总(中文优先) + 阈值记录英文
    fla_cn = g["fla"]
    fla_en = t.get("fla", "")
    note = fla_cn
    if fla_en and fla_en not in fla_cn:
        note = (note + "；" + fla_en).strip("；") if note else fla_en
    items.append(dict(
        en=g["en"], cas=g["cas"], cn=clean_cn(cn), cat=map_cat(g["cat"]),
        cat_raw=g["cat"], thr=fmt_thr(t.get("thr_val")), med=t.get("med"),
        note=note, pending=g["pending"],
        syn=[fc.normalize(x) for x in g["raw_ens"] if x.lower() != g["en"].lower()],
        src="发酵食品文献库",
    ))

json.dump(items, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("生成:", OUT, "唯一物质:", len(items))
print("  含阈值:", sum(1 for x in items if x["thr"]))
print("  有有效中文名:", sum(1 for x in items if x["cn"]))
print("  待核(无中文名):", sum(1 for x in items if x["pending"]))
