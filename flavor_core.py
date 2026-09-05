# -*- coding: utf-8 -*-
"""
风味物质自动检索 / 分类 / 阈值引擎
================================
- 内置风味化合物数据库（GC-MS 实测 64 种 + 常见风味化合物），含中英名、类别、
  气味阈值(μg/kg, 水相近似)、介质、备注、来源。
- normalize(): 名称归一化（大小写/空格/标点/连字符统一）。
- auto_classify(): 按命名中的官能团特征自动推断类别（未知物兜底）。
- match_compound(): 精确 / 包含 / 模糊 三级匹配，未命中则自动分类并提示补阈值。
- parse_gcms_excel(): 解析 Agilent MassHunter 导出的定量报告，提取化合物/RT/响应/响应比。
- enrich(): 对一批物质批量检索并合并原始量值。
用户可在 custom_compounds.json 中扩充数据库，重启即生效。
"""
import re
import os
import json
import difflib

HERE = os.path.dirname(os.path.abspath(__file__))

# ----------------------------------------------------------------------------
# 1. 名称归一化
# ----------------------------------------------------------------------------
def normalize(s):
    if not isinstance(s, str):
        return ''
    s = s.lower().replace('’', "'").replace('‘', "'")
    # 仅保留 字母 / 数字 / 中文，去掉空格、连字符、括号、逗号等
    s = re.sub(r'[^a-z0-9一-鿿]', '', s)
    return s


# ----------------------------------------------------------------------------
# 2. 官能团自动分类（未知物兜底）
# ----------------------------------------------------------------------------
TERPENE_HINTS = ['terpene', 'terpen', '萜', 'limonene', 'pinene', 'myrcene',
                 'ocimene', 'linalool', 'geraniol', 'citronellol', 'nerol',
                 'menthol', 'borneol', 'carveol', 'terpineol', 'caryophyllene',
                 'farnesol', 'bisabolol', 'camphene', 'terpinene', 'phellandrene',
                 'sabinene', 'carene', 'thujone', 'cedrol', 'fenchone', 'camphor']

def auto_classify(name):
    """按命名中的官能团特征自动推断类别（未知物兜底）。
    新增细分类别：内酯类、含硫化合物、吡嗪类、噻唑类/噻唑啉类、呋喃/呋喃酮类、
    含氮杂环/其他含氮化合物、萜烯类与含氧萜类。"""
    n = name.lower()
    # 1) 内酯：环状酯，单独成类
    if 'lactone' in n or '内酯' in n:
        return '内酯类'
    # 2) 吡嗪类（含氮杂环的子集，单独识别）
    if 'pyrazine' in n or '吡嗪' in n:
        return '吡嗪类'
    # 3) 噻唑类 / 噻唑啉类（含硫含氮杂环）
    if 'thiazole' in n or 'thiazoline' in n or '噻唑' in n:
        return '噻唑类 / 噻唑啉类'
    # 4) 含硫化合物：硫醇/硫醚/硫代等（注意：噻唑已在上面分流，此处不再捕获）
    if ('sulf' in n) or ('thiol' in n) or ('thia' in n) or ('mercapt' in n) \
       or ('disulfide' in n) or ('trisulfide' in n) or ('硫' in n) \
       or ('furfurylthiol' in n) or ('furanthiol' in n):
        return '含硫化合物'
    # 5) 呋喃 / 呋喃酮类（furfural/furfuryl 也是呋喃衍生物；furfurylthiol 已在前面的含硫规则分流）
    if any(h in n for h in ('furan', '呋喃', 'furaneol', 'furanone', 'furfural', 'furfuryl')):
        return '呋喃/呋喃酮类'
    # 6) 含氮杂环 / 其他含氮化合物（吡嗪、噻唑已单独分流）
    N_HETERO = ('pyridine', 'pyrrole', 'pyrrolidine', 'pyrimidine', 'indole',
                'quinoline', 'oxazole', 'isoxazole', 'pyrazole', 'imidazole',
                'azole', 'morpholine', 'piperidine', 'piperazine',
                '吡啶', '吡咯', '吲哚', '喹啉', '恶唑', '异恶唑', '吡唑',
                '咪唑', '氮杂', '吗啉', '哌啶', '哌嗪')
    if any(h in n for h in N_HETERO):
        return '含氮杂环/其他含氮化合物'
    # 7) 萜烯类与含氧萜类
    if any(h in n for h in TERPENE_HINTS):
        return '萜烯类与含氧萜类'
    if ('aldehyde' in n) or ('醛' in n) or re.search(r'anal$', n) or re.search(r'\bal$', n):
        return '醛类'
    if ('acid' in n) or ('oic' in n) or ('羧酸' in n) or ('酸' in n):
        return '酸类'
    if ('ester' in n) or ('ate' in n) or ('酯' in n):
        return '酯类'
    if ('ketone' in n) or ('酮' in n) or re.search(r'\bone$', n) \
       or 'diacetyl' in n or 'acetoin' in n:
        return '酮类'
    if ('alcohol' in n) or ('醇' in n) or re.search(r'(an|en|in|ol)ol$', n) or n.endswith('ol'):
        return '醇类'
    return '其他'


# ----------------------------------------------------------------------------
# 3. 风味化合物数据库
#    字段: en(英文规范名) cn(中文) cat(类别) thr(阈值) med(介质) note(备注) src(来源) syn(同义名)
#    阈值单位统一为 μg/kg(=ppb, 水相近似)，"≈" 表示量级近似；"—" 表示无文献值/无风味意义
# ----------------------------------------------------------------------------
COMPOUNDS = [
    # ---- 本批 GC-MS 实测 64 种 ----
    dict(en="Butanoic acid", cn="丁酸", cat="酸类", thr="≈240", med="水", note="奶酪/汗臭特征酸，奶油关键酸类风味物", src="Van Gemert 2011", syn=["Butyric acid"]),
    dict(en="Butanoic acid, methyl ester", cn="丁酸甲酯", cat="酯类", thr="≈70", med="水", note="果香/菠萝香，发酵及溶剂衍生酯", src="Van Gemert 2011"),
    dict(en="2-Heptanone", cn="2-庚酮", cat="酮类", thr="≈140", med="水", note="蓝纹奶酪典型霉香酮，阈值较低", src="Van Gemert 2011"),
    dict(en="Acetone", cn="丙酮", cat="酮类", thr="≈1.0×10⁵", med="水", note="阈值很高，弱风味；多为溶剂/代谢物", src="Van Gemert 2011"),
    dict(en="Hexanoic acid", cn="己酸", cat="酸类", thr="≈3.0×10³", med="水", note="脂肪酸臭/汗臭，奶酪风味贡献", src="Van Gemert 2011"),
    dict(en="Methyl Alcohol", cn="甲醇", cat="醇类", thr="≈1.0×10⁵", med="水", note="阈值很高，无风味意义（溶剂/杂质）", src="Van Gemert 2011", syn=["Methanol"]),
    dict(en="n-Hexane", cn="正己烷", cat="其他", thr="—", med="—", note="无风味意义，溶剂残留/前处理带入", src="—"),
    dict(en="Cyclotrisiloxane, hexamethyl-", cn="六甲基环三硅氧烷", cat="其他", thr="—", med="—", note="无风味意义，色谱柱流失/隔垫硅氧烷", src="—"),
    dict(en="Cyclotetrasiloxane, octamethyl-", cn="八甲基环四硅氧烷", cat="其他", thr="—", med="—", note="无风味意义，色谱柱流失/隔垫硅氧烷", src="—"),
    dict(en="Octanoic acid", cn="辛酸", cat="酸类", thr="≈3.0×10³", med="水", note="脂肪酸臭，奶酪/奶油风味", src="Van Gemert 2011", syn=["Caprylic acid"]),
    dict(en="Cyclohexasiloxane, dodecamethyl-", cn="十二甲基环六硅氧烷", cat="其他", thr="—", med="—", note="无风味意义，色谱柱流失/隔垫硅氧烷", src="—"),
    dict(en="2-Nonanone", cn="2-壬酮", cat="酮类", thr="≈3", med="水", note="霉香/酮香，阈值极低（奶油关键酮）", src="Van Gemert 2011"),
    dict(en="Cyclopentasiloxane, decamethyl-", cn="十甲基环五硅氧烷", cat="其他", thr="—", med="—", note="无风味意义，色谱柱流失/隔垫硅氧烷", src="—"),
    dict(en="Toluene", cn="甲苯", cat="其他", thr="≈500", med="水", note="溶剂/污染物，无正面风味贡献", src="Van Gemert 2011"),
    dict(en="Cyclooctasiloxane, hexadecamethyl-", cn="十六甲基环八硅氧烷", cat="其他", thr="—", med="—", note="无风味意义，色谱柱流失/隔垫硅氧烷", src="—"),
    dict(en="2-Undecanone", cn="2-十一酮", cat="酮类", thr="≈2", med="水", note="柑橘/霉香，阈值很低", src="Van Gemert 2011"),
    dict(en="2H-Pyran-2-one, tetrahydro-6-pentyl-", cn="γ-壬内酯", cat="酯类", thr="7–65", med="水", note="椰香/奶油香，关键内酯（环状酯）", src="SJTU风味库/Van Gemert", syn=["Gamma-nonalactone"]),
    dict(en="n-Decanoic acid", cn="正癸酸", cat="酸类", thr="≈1.0×10⁴", med="水", note="脂肪酸/蜡质气味", src="Van Gemert 2011", syn=["Capric acid"]),
    dict(en="Cyclobutanone, 2,3,3-trimethyl-", cn="2,3,3-三甲基环丁酮", cat="酮类", thr="—", med="—", note="无文献阈值；多为热处理加工衍生", src="—"),
    dict(en="Quinoline, 1,2-dihydro-2,2,4-trimethyl-", cn="2,2,4-三甲基-1,2-二氢喹啉", cat="其他", thr="—", med="—", note="可能来自橡胶抗氧化剂迁移，无风味意义", src="—"),
    dict(en="Silanol, trimethyl-", cn="三甲基硅醇", cat="其他", thr="—", med="—", note="无风味意义，硅烷化/柱流失产物", src="—"),
    dict(en="2H-Pyran-2-one, tetrahydro-6-propyl-", cn="γ-庚内酯", cat="酯类", thr="50–200", med="水", note="椰香/奶油香，短链内酯", src="Van Gemert 2011", syn=["Gamma-heptalactone"]),
    dict(en="Hexane, 3,3-dimethyl-", cn="3,3-二甲基己烷", cat="其他", thr="—", med="—", note="无风味意义，烷烃杂质", src="—"),
    dict(en="Cyclononasiloxane, octadecamethyl-", cn="十八甲基环九硅氧烷", cat="其他", thr="—", med="—", note="无风味意义，色谱柱流失/隔垫硅氧烷", src="—"),
    dict(en="Diazene, dimethyl-", cn="二甲基二氮烯", cat="其他", thr="—", med="—", note="无风味意义（可能假阳性/衍生）", src="—"),
    dict(en="2-Butanone", cn="2-丁酮（甲基乙基酮）", cat="酮类", thr="≈5.0×10³", med="水", note="溶剂/发酵酮，弱风味", src="Van Gemert 2011", syn=["Methyl ethyl ketone"]),
    dict(en="Dimethyl sulfone", cn="二甲基砜", cat="其他", thr="—", med="—", note="无风味意义，来源不明", src="—"),
    dict(en="Ethoxyacetylene", cn="乙氧基乙炔", cat="其他", thr="—", med="—", note="无风味意义（疑似假阳性）", src="—"),
    dict(en="2-Hexanone", cn="2-己酮", cat="酮类", thr="400–800", med="水", note="果香/青香酮", src="Van Gemert 2011"),
    dict(en="1,2-Ethanediol, 1,2-diphenyl-, [R-(R*,R*)]-", cn="1,2-二苯基-1,2-乙二醇（氢化苯偶姻）", cat="其他", thr="—", med="—", note="无风味意义，可能为污染物", src="—"),
    dict(en="Carbamic acid, methyl-, 3-methylphenyl ester", cn="3-甲基苯基甲基氨基甲酸酯", cat="酯类", thr="—", med="—", note="无风味意义（疑似污染物/农药相关）", src="—"),
    dict(en="3,4-Dihydroxybutan-2-one, O,O'-diacetyl", cn="O,O'-二乙酰基-3,4-二羟基丁-2-酮", cat="酮类", thr="—", med="—", note="无文献阈值，疑似衍生/假阳性", src="—"),
    dict(en="Benzene, 1,3-dimethyl-", cn="1,3-二甲基苯（间二甲苯）", cat="其他", thr="50–100", med="水", note="溶剂/污染物，无正面风味贡献", src="Van Gemert 2011"),
    dict(en="2-Octanone", cn="2-辛酮", cat="酮类", thr="50–200", med="水", note="果香/霉香酮", src="Van Gemert 2011"),
    dict(en="Benzaldehyde", cn="苯甲醛", cat="醛类", thr="≈350", med="水", note="杏仁香，关键醛类风味物", src="Van Gemert 2011"),
    dict(en="2,4-Dihydroxybenzaldehyde, 2TMS derivative", cn="2,4-二羟基苯甲醛（2TMS 衍生）", cat="醛类", thr="—（母体≈1.0×10³）", med="水", note="来自酚类/糖苷，TMS 衍生化形式", src="—"),
    dict(en="1-Hexanol, 2-ethyl-", cn="2-乙基-1-己醇", cat="醇类", thr="≈1.0×10³", med="水", note="玫瑰/青香；亦与塑化剂相关", src="Van Gemert 2011", syn=["2-Ethylhexanol"]),
    dict(en="4-Octene, (E)-", cn="(E)-4-辛烯", cat="其他", thr="—", med="—", note="无风味意义（烯烃杂质）", src="—"),
    dict(en="3-Heptene, 4-methyl-", cn="4-甲基-3-庚烯", cat="其他", thr="—", med="—", note="无风味意义（烯烃杂质）", src="—"),
    dict(en="1-Pentanol", cn="1-戊醇", cat="醇类", thr="≈4.0×10³", med="水", note="醇香/麦芽香", src="Van Gemert 2011"),
    dict(en="Acetic acid, TBDMS derivative", cn="乙酸 TBDMS 衍生物", cat="酸类", thr="—", med="—", note="乙酸的硅烷化衍生形式", src="—"),
    dict(en="Azetidine", cn="氮杂环丁烷", cat="其他", thr="—", med="—", note="无风味意义（含氮杂环）", src="—"),
    dict(en="Acetoin", cn="乙偶姻（3-羟基-2-丁酮）", cat="酮类", thr="≈5.0×10⁴", med="水", note="黄油/奶油香，但阈值较高", src="Van Gemert 2011"),
    dict(en="Benzaldehyde, 3,4-dimethyl-", cn="3,4-二甲基苯甲醛", cat="醛类", thr="—", med="—", note="芳香醛，无报道阈值", src="—"),
    dict(en="Benzaldehyde, 3-ethyl-", cn="3-乙基苯甲醛", cat="醛类", thr="—", med="—", note="芳香醛，无报道阈值", src="—"),
    dict(en="Acetic acid", cn="乙酸", cat="酸类", thr="≈1.0×10⁵", med="水", note="醋香，阈值较高", src="Van Gemert 2011", syn=["Ethanoic acid"]),
    dict(en="Pentane, 2,3,3,4-tetramethyl-", cn="2,3,3,4-四甲基戊烷", cat="其他", thr="—", med="—", note="无风味意义（烷烃杂质）", src="—"),
    dict(en="Heptane, 3,3,4-trimethyl-", cn="3,3,4-三甲基庚烷", cat="其他", thr="—", med="—", note="无风味意义（烷烃杂质）", src="—"),
    dict(en="Pentane, 3-ethyl-", cn="3-乙基戊烷", cat="其他", thr="—", med="—", note="无风味意义（烷烃杂质）", src="—"),
    dict(en="Caprolactam", cn="己内酰胺", cat="其他", thr="—", med="—", note="无风味意义，可能来自包装/尼龙迁移", src="—"),
    dict(en="Acetic acid, 2-(dimethylamino)ethyl ester", cn="乙酸 2-(二甲氨基)乙酯", cat="酯类", thr="—", med="—", note="无报道阈值", src="—"),
    dict(en="Pentanal, 2-methyl-", cn="2-甲基戊醛", cat="醛类", thr="1–10", med="水", note="麦芽/可可香，阈值低", src="Van Gemert 2011"),
    dict(en="Undecane, 2,7-dimethyl-", cn="2,7-二甲基十一烷", cat="其他", thr="—", med="—", note="无风味意义（烷烃杂质）", src="—"),
    dict(en="1-Dodecanol, 3,7,11-trimethyl-", cn="3,7,11-三甲基-1-十二醇（法尼醇型倍半萜醇）", cat="萜烯类", thr="—", med="—", note="萜醇（倍半萜），痕量，无报道阈值", src="—"),
    dict(en="1-Hexanone, 5-methyl-1-phenyl-", cn="5-甲基-1-苯基-1-己酮", cat="酮类", thr="—", med="—", note="芳香酮，无报道阈值", src="—"),
    dict(en="3-Pentanamine", cn="3-戊胺", cat="其他", thr="—", med="—", note="无风味意义（胺类）", src="—"),
    dict(en="Benzoic acid, hydrazide", cn="苯甲酰肼", cat="其他", thr="—", med="—", note="无风味意义（酰肼）", src="—"),
    dict(en="2-tert-Butyl-6-methyl-phenol, acetate", cn="乙酸 2-叔丁基-6-甲基苯酯", cat="酯类", thr="—", med="—", note="可能来自抗氧化剂 BHT 衍生", src="—"),
    dict(en="3-甲基庚烷", cn="3-甲基庚烷", cat="其他", thr="—", med="—", note="无风味意义（烷烃杂质）", src="—"),
    dict(en="Bromochloronitromethane", cn="溴氯硝基甲烷", cat="其他", thr="—", med="—", note="无风味意义（消毒副产物/污染物）", src="—"),
    dict(en="1,4-Benzenediol, 2,5-bis(1,1-dimethylethyl)-", cn="2,5-二叔丁基-1,4-苯二酚", cat="其他", thr="—", med="—", note="BHT 氧化产物（抗氧化剂来源）", src="—"),
    dict(en="Cyclopentane, 1,1-dimethyl-", cn="1,1-二甲基环戊烷", cat="其他", thr="—", med="—", note="无风味意义（环烷烃）", src="—"),
    dict(en="Cinnamaldehyde, (E)-", cn="(E)-肉桂醛", cat="醛类", thr="≈1.0×10³", med="水", note="肉桂香，关键醛类风味物", src="Van Gemert 2011", syn=["Cinnamaldehyde"]),
    dict(en="Benzothiazole", cn="苯并噻唑", cat="其他", thr="—", med="—", note="无风味意义（含氮/硫杂环，污染源）", src="—"),

    # ---- 常见风味化合物（扩充库，阈值均为水相近似文献值） ----
    dict(en="Acetaldehyde", cn="乙醛", cat="醛类", thr="≈15", med="水", note="青香/氧化味", src="Van Gemert 2011"),
    dict(en="Hexanal", cn="己醛", cat="醛类", thr="≈4.5", med="水", note="青草/脂肪氧化香", src="Van Gemert 2011"),
    dict(en="(E)-2-Hexenal", cn="反-2-己烯醛（叶醛）", cat="醛类", thr="≈17", med="水", note="青叶香", src="Van Gemert 2011"),
    dict(en="Heptanal", cn="庚醛", cat="醛类", thr="≈3", med="水", note="脂肪/柑橘香", src="Van Gemert 2011"),
    dict(en="Octanal", cn="辛醛", cat="醛类", thr="≈0.7", med="水", note="柑橘/脂肪香", src="Van Gemert 2011"),
    dict(en="Nonanal", cn="壬醛", cat="醛类", thr="≈1", med="水", note="柑橘/玫瑰/油脂香", src="Van Gemert 2011"),
    dict(en="Decanal", cn="癸醛", cat="醛类", thr="≈2", med="水", note="甜橙/脂肪香", src="Van Gemert 2011"),
    dict(en="(E)-2-Nonenal", cn="反-2-壬烯醛", cat="醛类", thr="≈0.08", med="水", note="黄瓜/老化油脂味，阈值极低", src="Van Gemert 2011"),
    dict(en="(E,E)-2,4-Decadienal", cn="反,反-2,4-癸二烯醛", cat="醛类", thr="≈0.07", med="水", note="油炸/鸡脂香，阈值极低", src="Van Gemert 2011"),
    dict(en="Methional", cn="3-甲硫基丙醛", cat="醛类", thr="≈0.2", med="水", note="煮土豆/肉汤香", src="Van Gemert 2011"),
    dict(en="Furfural", cn="糠醛", cat="醛类", thr="≈3.0×10³", med="水", note="焦糖/杏仁香", src="Van Gemert 2011"),
    dict(en="Diacetyl", cn="丁二酮（双乙酰）", cat="酮类", thr="≈2.5", med="水", note="黄油/奶油关键酮，阈值低", src="Van Gemert 2011", syn=["2,3-Butanedione"]),
    dict(en="2,3-Pentanedione", cn="2,3-戊二酮", cat="酮类", thr="≈2", med="水", note="奶油/黄油香", src="Van Gemert 2011"),
    dict(en="3-Hydroxy-2-butanone", cn="3-羟基-2-丁酮", cat="酮类", thr="≈5.0×10⁴", med="水", note="黄油香（即乙偶姻）", src="Van Gemert 2011"),
    dict(en="Ethyl butyrate", cn="丁酸乙酯", cat="酯类", thr="≈0.1", med="水", note="菠萝/果香，阈值极低", src="Van Gemert 2011"),
    dict(en="Ethyl hexanoate", cn="己酸乙酯", cat="酯类", thr="≈1", med="水", note="果香/酒香", src="Van Gemert 2011"),
    dict(en="Isoamyl acetate", cn="乙酸异戊酯", cat="酯类", thr="≈2", med="水", note="香蕉香", src="Van Gemert 2011"),
    dict(en="Ethyl acetate", cn="乙酸乙酯", cat="酯类", thr="≈5.0×10³", med="水", note="果香/溶剂香，阈值高", src="Van Gemert 2011"),
    dict(en="Methyl butyrate", cn="丁酸甲酯", cat="酯类", thr="≈70", med="水", note="果香/菠萝香", src="Van Gemert 2011"),
    dict(en="Limonene", cn="柠檬烯", cat="萜烯类", thr="≈200", med="水", note="柑橘香（萜烯）", src="Van Gemert 2011"),
    dict(en="alpha-Pinene", cn="α-蒎烯", cat="萜烯类", thr="≈6", med="水", note="松木香（萜烯）", src="Van Gemert 2011"),
    dict(en="Myrcene", cn="月桂烯", cat="萜烯类", thr="≈15", med="水", note="青草/胡椒香（萜烯）", src="Van Gemert 2011"),
    dict(en="Linalool", cn="芳樟醇", cat="萜烯类", thr="≈6", med="水", note="花香/薰衣草（萜醇）", src="Van Gemert 2011"),
    dict(en="Geraniol", cn="香叶醇", cat="萜烯类", thr="≈10", med="水", note="玫瑰香（萜醇）", src="Van Gemert 2011"),
    dict(en="Citronellol", cn="香茅醇", cat="萜烯类", thr="≈20", med="水", note="玫瑰/柠檬香（萜醇）", src="Van Gemert 2011"),
    dict(en="Menthol", cn="薄荷醇", cat="萜烯类", thr="≈2", med="水", note="薄荷凉感（萜醇）", src="Van Gemert 2011"),
    dict(en="Vanillin", cn="香兰素", cat="其他", thr="≈20", med="水", note="香草香（酚醛类）", src="Van Gemert 2011"),
    dict(en="Ethyl maltol", cn="乙基麦芽酚", cat="其他", thr="≈4", med="水", note="焦糖/棉花糖香", src="Van Gemert 2011"),
    dict(en="Maltol", cn="麦芽酚", cat="其他", thr="≈35", med="水", note="焦糖/面包香", src="Van Gemert 2011"),
    dict(en="gamma-Decalactone", cn="γ-癸内酯", cat="酯类", thr="≈11", med="水", note="桃香内酯", src="Van Gemert 2011"),
    dict(en="delta-Decalactone", cn="δ-癸内酯", cat="酯类", thr="≈100", med="水", note="奶油/桃香内酯", src="Van Gemert 2011"),
    dict(en="guaiacol", cn="愈创木酚", cat="其他", thr="≈10", med="水", note="烟熏/酚香", src="Van Gemert 2011"),
    dict(en="4-Vinylguaiacol", cn="4-乙烯基愈创木酚", cat="其他", thr="≈3", med="水", note="丁香/烟熏香", src="Van Gemert 2011"),
    dict(en="p-Cresol", cn="对甲酚", cat="其他", thr="≈0.05", med="水", note="马厩/动物臭，阈值极低", src="Van Gemert 2011"),
    dict(en="Skatole", cn="粪臭素（3-甲基吲哚）", cat="其他", thr="≈0.05", med="水", note="粪便/灵猫香，阈值极低", src="Van Gemert 2011", syn=["3-Methylindole"]),
    dict(en="Indole", cn="吲哚", cat="其他", thr="≈140", med="水", note="高浓度粪臭、低浓度花香", src="Van Gemert 2011"),
    dict(en="Dimethyl sulfide", cn="二甲基硫醚", cat="其他", thr="≈0.3", med="水", note="煮玉米/海产香，阈值低", src="Van Gemert 2011"),
    dict(en="Dimethyl trisulfide", cn="二甲基三硫", cat="其他", thr="≈0.01", med="水", note="洋葱/蒜臭，阈值极低", src="Van Gemert 2011"),
    dict(en="Methanethiol", cn="甲硫醇", cat="其他", thr="≈0.02", med="水", note="腐败/硫化物臭", src="Van Gemert 2011"),
    dict(en="2-Acetyl-1-pyrroline", cn="2-乙酰基-1-吡咯啉", cat="其他", thr="≈0.0001", med="水", note="爆米花/米饭香，阈值极低", src="Van Gemert 2011"),
    dict(en="2-Acetylpyrazine", cn="2-乙酰基吡嗪", cat="其他", thr="≈50", med="水", note="坚果/烘烤香", src="Van Gemert 2011"),
    dict(en="Phenylethyl alcohol", cn="苯乙醇", cat="醇类", thr="≈1.4×10³", med="水", note="玫瑰/蜂蜜香", src="Van Gemert 2011", syn=["2-Phenylethanol"]),
    dict(en="Isoamyl alcohol", cn="异戊醇", cat="醇类", thr="≈1.0×10³", med="水", note="麦芽/杂醇油香", src="Van Gemert 2011"),
    dict(en="1-Propanol", cn="1-丙醇", cat="醇类", thr="≈4.0×10³", med="水", note="醇香/杂醇", src="Van Gemert 2011"),
    dict(en="Ethanol", cn="乙醇", cat="醇类", thr="≈1.0×10⁵", med="水", note="酒香，阈值高", src="Van Gemert 2011"),
    dict(en="1-Butanol", cn="1-丁醇", cat="醇类", thr="≈5.0×10³", med="水", note="醇香/溶剂", src="Van Gemert 2011"),
    dict(en="1-Hexanol", cn="1-己醇", cat="醇类", thr="≈2.5×10³", med="水", note="青草/脂肪醇香", src="Van Gemert 2011"),
    dict(en="1-Octanol", cn="1-辛醇", cat="醇类", thr="≈1.1×10³", med="水", note=" citrus/脂肪醇香；常作内标", src="Van Gemert 2011"),
    dict(en="Lactic acid", cn="乳酸", cat="酸类", thr="≈1.0×10⁵", med="水", note="酸味，气味弱", src="Van Gemert 2011"),
    dict(en="Propionic acid", cn="丙酸", cat="酸类", thr="≈3.0×10³", med="水", note="刺激性酸臭", src="Van Gemert 2011"),
    dict(en="Valeric acid", cn="戊酸", cat="酸类", thr="≈700", med="水", note="腐败/汗酸臭", src="Van Gemert 2011"),
    dict(en="3-Methylbutanoic acid", cn="3-甲基丁酸（异戊酸）", cat="酸类", thr="≈50", med="水", note="刺激性酸臭/奶酪", src="Van Gemert 2011"),
    dict(en="2-Methylbutanoic acid", cn="2-甲基丁酸", cat="酸类", thr="≈60", med="水", note="奶酪/汗酸", src="Van Gemert 2011"),
]


# ----------------------------------------------------------------------------
# 4. 建立检索索引（英文规范名 + 同义名 + 中文名）
#    同一归一化键若已存在则【不覆盖】，保证精选条目优先匹配。
# ----------------------------------------------------------------------------
def _build_index():
    idx = {}
    for c in COMPOUNDS:
        for key in [c['en']] + c.get('syn', []) + [c.get('cn', '')]:
            nk = normalize(key)
            # 跳过退化键：空 / 纯数字(如 '2') / 单字符(如 '醇')，
            # 否则「包含匹配」会让这些键成为大量查询的子串而批量误命中
            if not nk or nk.isdigit() or len(nk) < 2:
                continue
            idx.setdefault(nk, c)
    return idx


def _load_custom():
    """从 custom_compounds.json 合并用户自定义化合物（覆盖同名）。"""
    p = os.path.join(HERE, 'custom_compounds.json')
    if not os.path.exists(p):
        return
    try:
        with open(p, 'r', encoding='utf-8') as f:
            extra = json.load(f)
        for c in extra:
            COMPOUNDS.append(c)
    except Exception as e:
        print('[warn] 读取 custom_compounds.json 失败:', e)


# ----------------------------------------------------------------------------
# FTDB 之外、精选库仍缺 CAS 的物质：经核实的标准 CAS（公共化学知识 / PubChem）。
# 仅用于【补充】CAS，绝不覆盖阈值/备注/类别。键为 normalize(英文名)。
# 取代物一律用其【自身】CAS，不用母体 CAS（如苯甲醛取代物用 5271-06-1，非 100-52-7）。
# ----------------------------------------------------------------------------
CAS_OVERRIDE = {
    normalize('Butanoic acid'): '107-92-6',
    normalize('1-Pentanol'): '71-41-0',
    normalize('Acetoin'): '513-86-0',
    normalize('Diacetyl'): '431-03-8',
    normalize('3-Hydroxy-2-butanone'): '513-86-0',
    normalize('Limonene'): '138-86-3',
    normalize('Myrcene'): '123-35-3',
    normalize('Linalool'): '78-70-6',
    normalize('Geraniol'): '106-24-1',
    normalize('Citronellol'): '106-22-9',
    normalize('Skatole'): '83-34-1',
    normalize('1-Propanol'): '71-23-8',
    normalize('1-Butanol'): '71-36-3',
    normalize('1-Hexanol'): '111-27-3',
    normalize('1-Octanol'): '111-87-5',
    normalize('3-Methylbutanoic acid'): '503-74-2',
    normalize('2-Butanone'): '78-93-3',
    normalize('n-Decanoic acid'): '334-48-5',
    normalize('1-Hexanol, 2-ethyl-'): '104-76-7',
    normalize('(E)-2-Hexenal'): '6728-26-3',
    normalize('alpha-Pinene'): '80-56-8',
    normalize('Ethyl maltol'): '4940-11-2',
    normalize('4-Vinylguaiacol'): '7786-61-0',
    normalize('Dimethyl trisulfide'): '3658-80-8',
    normalize('Benzaldehyde, 3,4-dimethyl-'): '5271-06-1',
    normalize('Benzaldehyde, 3-ethyl-'): '34246-54-3',
    normalize('Quinoline, 1,2-dihydro-2,2,4-trimethyl-'): '147-47-7',
    normalize('Benzoic acid, hydrazide'): '613-94-5',
    normalize('1,4-Benzenediol, 2,5-bis(1,1-dimethylethyl)-'): '88-58-4',
    normalize('Acetic acid, 2-(dimethylamino)ethyl ester'): '625-56-9',
    normalize('3-Pentanamine'): '616-24-0',
    normalize('Dimethyl sulfone'): '67-71-0',
    normalize('Silanol, trimethyl-'): '1066-40-6',
    normalize('Ethoxyacetylene'): '689-97-4',
    normalize('Cyclotrisiloxane, hexamethyl-'): '541-05-9',
    normalize('Cyclotetrasiloxane, octamethyl-'): '556-67-2',
    normalize('Cyclopentasiloxane, decamethyl-'): '541-02-6',
    normalize('Cyclohexasiloxane, dodecamethyl-'): '540-97-6',
    normalize('Cyclooctasiloxane, hexadecamethyl-'): '556-71-8',
    normalize('Cyclononasiloxane, octadecamethyl-'): '556-72-9',
    normalize('n-Hexane'): '110-54-3',
    normalize('Hexane, 3,3-dimethyl-'): '562-49-2',
    normalize('2-tert-Butyl-6-methyl-phenol, acetate'): '2219-82-1',
    normalize('2,4-Dihydroxybenzaldehyde, 2TMS derivative'): '95-01-2',
    normalize('Acetic acid, TBDMS derivative'): '64-19-7',
    normalize('Carbamic acid, methyl-, 3-methylphenyl ester'): '1123-72-4',
    normalize('Diazene, dimethyl-'): '821-12-5',
    normalize('Cyclopentane, 1,1-dimethyl-'): '1638-26-2',
    normalize('1,2-Ethanediol, 1,2-diphenyl-, [R-(R*,R*)]-'): '492-70-6',
    normalize('Benzene, 1,3-dimethyl-'): '108-38-3',   # 间二甲苯 m-xylene
}


def _load_ftdb():
    """并入 FlavorThresholdDB (XQplayer/HXQLab v1.5.0) 的 aroma_data_merged.json。
    - 精选条目优先：仅【补充】CAS / 风味类别，不覆盖其阈值与备注；
    - 其余 FTDB 化合物追加进库；DB 索引不覆盖已存在的精选键。"""
    p = os.path.join(HERE, 'ftdb_compounds.json')
    if not os.path.exists(p):
        return
    try:
        data = json.load(open(p, 'r', encoding='utf-8'))
    except Exception as e:
        print('[warn] 读取 ftdb_compounds.json 失败:', e)
        return
    # 建立 FTDB 查找表：按英文全名 / 括号片段 / 逗号前段 / 中文名 / 同义名（精确匹配，杜绝母体错配）
    fmap = {}
    for r in data:
        segs = set()
        en = r.get('en') or ''
        segs.add(normalize(en))
        for part in re.findall(r'\(([^)]*)\)', en):      # 括号同义名（如 butyric acid）
            segs.add(normalize(part))
        segs.add(normalize(re.sub(r'[,(].*', '', en)))    # 逗号/括号前的主名段
        if r.get('cn'):
            segs.add(normalize(r['cn']))
        for s in r.get('syn', []) or []:
            segs.add(normalize(s))
        for s in segs:
            if s:
                fmap.setdefault(s, r)
    # 1) 富精选条目：补 CAS（FTDB 精确匹配 → 人工核实覆盖表）
    curated_src = ('Van Gemert 2011', '—', 'SJTU风味库/Van Gemert', '自定义')
    for c in COMPOUNDS:
        if c.get('cas') or c.get('src') not in curated_src:
            continue
        m = (fmap.get(normalize(c['en']))
             or (c.get('cn') and fmap.get(normalize(c['cn'])))
             or next((fmap.get(normalize(s)) for s in c.get('syn', []) if fmap.get(normalize(s))), None))
        if m and m.get('cas'):
            c['cas'] = m['cas']
            continue
        ov = CAS_OVERRIDE.get(normalize(c['en'])) or (c.get('cn') and CAS_OVERRIDE.get(normalize(c['cn'])))
        if ov:
            c['cas'] = ov
    # 2) 追加 FTDB 新条目（后续 _build_index 已存在键不覆盖）
    COMPOUNDS.extend(data)
    # 3) 去重 COMPOUNDS 列表（同一归一化英文名仅保留首次出现=精选优先）
    seen = set(); dedupe = []
    for c in COMPOUNDS:
        k = normalize(c['en'])
        if k in seen:
            continue
        seen.add(k); dedupe.append(c)
    COMPOUNDS[:] = dedupe


def _drop_column_bleed():
    """剔除 GC 色谱柱流失相关物质（聚二甲基硅氧烷固定相 bleed 产生的硅氧烷类）。
    仅从运行库排除，保留原始 curated 定义以便追溯。"""
    pat = re.compile(r'siloxane|silanol|硅氧烷|硅醇|polysiloxane', re.I)
    dropped = [c for c in COMPOUNDS if pat.search(' '.join([str(c.get('en', '')), str(c.get('cn', '')), str(c.get('note', ''))]))]
    if dropped:
        names = [c['en'] for c in dropped]
        COMPOUNDS[:] = [c for c in COMPOUNDS if c not in dropped]
        print('[info] 已剔除色谱柱流失(硅氧烷)相关物质 %d 种: %s' % (len(dropped), ', '.join(names)))


def clean_note(note):
    """清洗备注：删除来源/程序类标注（大论文、图片风味描述、原thr、FTDB程序备注等），
    仅保留真正的风味词。"""
    if not note:
        return '—'
    s = note
    # 1) 大论文X(...) ：只删"大论文X"标签，保留半角括号内的中文风味词
    s = re.sub(r'大论文[A-Za-z]?\d*\s*\(([^)]*)\)', r'\1', s)
    # 2) 大论文X：/大论文X: 标签（删标签，保留其后中文风味词）
    s = re.sub(r'大论文[A-Za-z]?\d*\s*[：:]', '', s)
    # 3) 大论文X 独立标签（如 大论文D1 / 大论文C2）
    s = re.sub(r'大论文[A-Za-z]?\d+', '', s)
    # 4) 图片风味描述: 标签前缀（删前缀，保留其后内容）
    s = re.sub(r'图片风味描述\s*[:：]\s*', '', s)
    # 5) ；原thr=xxx 溯源标签
    s = re.sub(r'[；;]?\s*原thr=[^；;]*', '', s)
    # 6) 全角括号说明（FTDB无标注/空气阈值/量纲/别名）
    s = re.sub(r'（[^）]*）', '', s)
    # 7) 半角括号说明 (β-Damascenone) (= Aldol) (γ-...) (含硫) 等
    s = re.sub(r'\([^)]*\)', '', s)
    # 8) 同 xxx； 别名前缀  与  ；注：... 说明
    s = re.sub(r'同\s+[^；;]+[；;]', '', s)
    s = re.sub(r'[；;]?\s*注：.*$', '', s)
    # 9) 来源/占位标注
    s = re.sub(r'HS-GC-MS\s*标准品', '', s)
    s = re.sub(r'Odor描述查询不到', '', s)
    s = re.sub(r'Odor查询不到', '', s)
    s = re.sub(r'已在上面', '', s)
    # 10) 纯结构说明（无风味词）
    s = re.sub(r'[；;]?\s*表中标反式[^；;]*', '', s)
    s = re.sub(r'[；;]?\s*实际为顺式', '', s)
    # 规范化分隔符
    s = re.sub(r'[；;]\s*', '；', s)
    s = re.sub(r'；+', '；', s)
    s = s.strip('；').strip()
    # 去除开头的占位符 "—；" 或单纯的 "—"
    s = re.sub(r'^—[；;]?\s*', '', s).strip('；').strip()
    return s if s else '—'


def _clean_notes():
    """清洗全部条目的 note：删除来源/程序标注，仅保留风味词。"""
    n_changed = 0
    for c in COMPOUNDS:
        raw = c.get('note')
        cl = clean_note(raw)
        if cl != raw:
            c['note'] = cl
            n_changed += 1
    if n_changed:
        print('[info] 清洗备注(note): %d 条已清理，仅保留风味词' % n_changed)


# 各类别的典型呈香/风味描述（FTDB 无描述时兜底显示，避免备注空白）
CAT_FLAVOR = {
    '醛类': '青香、脂肪香、果香；阈值低者呈香贡献大（如己醛呈青草脂肪味）',
    '酮类': '奶油香、霉香、果香；双乙酰/乙酰基酮呈典型黄油奶油香',
    '酯类': '果香、花香、甜香；多为发酵与水果特征香气（如乙酸乙酯果香、己酸乙酯酒香）',
    '酸类': '酸香、奶酪/发酵臭；短链酸呈刺激性酸味，长链酸呈奶酪/奶油脂肪感',
    '醇类': '醇香、青香、植物香；高级醇与萜醇呈玫瑰/薰衣草等花香',
    '萜烯类与含氧萜类': '柑橘香、松木香、花香；柠檬烯呈柑橘香，芳樟醇/香叶醇呈花香',
    '含氮杂环/其他含氮化合物': '烤香、焦香、坚果香；吡咯/吲哚等含氮杂环呈烘烤/动物香',
    '呋喃/呋喃酮类': '焦糖香、烘烤香、水果香；呋喃酮呈典型焦糖/菠萝香',
    '噻唑类 / 噻唑啉类': '肉香、烤香、坚果香；噻唑类呈烘烤/硫化物香',
    '其他': '杂环/硫化/酚类等；多为特征风味（烤肉香、葱蒜硫香、烟熏酚香）或溶剂/污染物',
}

def _fix_display():
    """显示兜底：① cn 无中文则回退英文全名；② note 为空的风味物质补类别香型描述。"""
    n_cn = 0
    n_note = 0
    for c in COMPOUNDS:
        cn = (c.get('cn') or '').strip()
        if cn and not re.search(r'[一-鿿]', cn):
            c['cn'] = c.get('en', cn)   # 英文碎片当中文名 -> 回退英文全名
            n_cn += 1
        note = (c.get('note') or '').strip()
        if note in ('', '—'):
            cat = c.get('cat') or '其他'
            c['note'] = CAT_FLAVOR.get(cat, CAT_FLAVOR['其他'])
            n_note += 1
    if n_cn:
        print('[info] 修正中文名(cn): %d 条英文碎片回退为英文全名' % n_cn)
    if n_note:
        print('[info] 备注兜底: %d 条无描述物质已补"类别香型"描述' % n_note)


def _clean_cn_references():
    """清理中文名里的前导参考文献标注，如「2- Hofmann &Schieberle(1997) 丙酰基-2-噻唑啉」
    只保留真正的中文名「丙酰基-2-噻唑啉」。

    识别规则：cn 含中文，但开头有一段非中文前缀，且该前缀像文献标注
    （含 4 位年份 / & 符号 / 字母后接括号 / 以「数字-」开头），则截掉该前缀，
    从首个中文字符起保留。中文名本身可能含数字/连字符/括号，不受影响。"""
    YEAR = re.compile(r'\d{4}')
    n = 0
    for c in COMPOUNDS:
        cn = c.get('cn') or ''
        if not re.search(r'[一-鿿]', cn):
            continue
        i = None
        for idx, ch in enumerate(cn):
            if '一' <= ch <= '鿿':
                i = idx
                break
        if i is None or i == 0:
            continue
        prefix = cn[:i]
        if not prefix.strip():
            continue
        if (YEAR.search(prefix) or '&' in prefix
                or re.search(r'[A-Za-z]\s*\(', prefix)):
            new = cn[i:].strip()
            if new and re.search(r'[一-鿿]', new):
                c['cn'] = new
                n += 1
    if n:
        print('[info] 清理中文名前导文献标注: %d 条（仅保留中文名）' % n)


def _normalize_thresholds_ugl():
    """全库阈值统一折算为 μg/L（之前仅 CSV 文件做了转换，FTDB / 发酵食品文献库等
    来源的阈值仍停留在原始文本：≈0.019、400–800、1.0×10⁵ 等，无单位且不可直接计算 OAV）。
    此处对每条化合物用 convert_threshold_ugl 折算：有单位按单位换算，无单位纯数字按 μg/L 处理，
    写入 thr(显示) / thr_num(数值) / thr_raw(原始文本)。"""
    n = 0
    for c in COMPOUNDS:
        raw = (c.get('thr_raw') or c.get('thr') or '').strip()
        if not raw or raw in ('—', '-', '–'):
            if c.get('thr') != '—' or c.get('thr_num') is not None:
                c['thr'] = '—'
                c['thr_num'] = None
                n += 1
            c.setdefault('thr_raw', '')
            continue
        thr, num, med, _st = convert_threshold_ugl(raw)
        c['thr_raw'] = (c.get('thr_raw') or raw).strip()
        if c.get('thr') != thr or c.get('thr_num') != num:
            n += 1
        c['thr'] = thr
        c['thr_num'] = num
        if med and med != '—' and not (c.get('med') or '').strip():
            c['med'] = med
    if n:
        print('[info] 全库阈值统一折算 μg/L: %d 条阈值/数值更新' % n)


def _clean_cn_spaces():
    """清理中文名中 Excel 解析残留的多余空格（断词），如「3-  甲基」→「3-甲基」。
    只压缩词内连续 2 个以上空格，不影响「乙酸 2-(二甲氨基)乙酯」这类合法单空格。"""
    n = 0
    for c in COMPOUNDS:
        cn = c.get('cn') or ''
        if not cn:
            continue
        new = re.sub(r'(?<=\S)\s{2,}(?=\S)', '', cn)   # 词内断词空格
        new = re.sub(r'\s{2,}', ' ', new).strip()         # 其余折叠为单空格
        if new != cn:
            c['cn'] = new
            n += 1
    if n:
        print('[info] 清理中文名多余空格(断词): %d 条' % n)


def _fix_pyrazine_cn():
    """吡嗪类中文名系统校正：源库中大量吡嗪条目被错填为笼统的「甲基吡嗪/乙基甲基吡嗪」等。
    此处根据英文名系统生成规范中文名（位置-取代基-吡嗪），覆盖烷基/烷氧基/乙酰基/硫基/苯基/乙烯基等。
    仅对「英文名可被完整解析」的条目生效；含截断(空格-)、环并杂环、文献箭头等异常的直接跳过，保留原值。"""
    PYR_REPL = [
        ('TRIMETHYL', '三甲基'), ('TETRAMETHYL', '四甲基'), ('DIMETHYL', '二甲基'),
        ('TRIETHYL', '三乙基'), ('DIETHYL', '二乙基'),
        ('METHYLTHIO', '甲硫基'), ('ETHYLTHIO', '乙硫基'), ('PHENYLTHIO', '苯硫基'),
        ('ISOPROPYL', '异丙基'), ('ISOBUTYL', '异丁基'), ('ISOAMYL', '异戊基'),
        ('ISOPENTYL', '异戊基'), ('ISOHEXYL', '异己基'), ('SEC-BUTYL', '仲丁基'),
        ('METHOXY', '甲氧基'), ('ETHOXY', '乙氧基'), ('PHENOXY', '苯氧基'),
        ('ACETYL', '乙酰基'), ('VINYL', '乙烯基'), ('ETHENYL', '乙烯基'),
        ('METHYL', '甲基'), ('ETHYL', '乙基'), ('PROPYL', '丙基'), ('BUTYL', '丁基'),
        ('PENTYL', '戊基'), ('HEXYL', '己基'), ('HEPTYL', '庚基'), ('OCTYL', '辛基'),
        ('DECYL', '癸基'), ('PYRAZINE', '吡嗪'),
    ]

    def _gen(en):
        if not en:
            return None
        if ' -' in en or en.rstrip().endswith('-') or '→' in en \
           or 'CYCLOPENTAN' in en.upper() or '[B]' in en:
            return None
        s = en.upper()
        for a, b in PYR_REPL:
            s = s.replace(a, b)
        # 仍有英文字母（除去立体化学 S/R/Z/E 与 +/−/括号）= 有未覆盖亚结构，跳过
        if re.search(r'[A-Za-z]', re.sub(r'[SRZE+\-()]', '', s)):
            return None
        return s

    n = 0
    for c in COMPOUNDS:
        if (c.get('cat') or '').strip() != '吡嗪类':
            continue
        new = _gen(c.get('en') or '')
        if new and new != (c.get('cn') or '').strip():
            c['cn'] = new
            n += 1
    if n:
        print('[info] 吡嗪类中文名系统校正: %d 条（按英文名→位置-取代基-吡嗪）' % n)


def _valid_cn(s):
    """有效中文名：含中文字符 且 不是 (无中文名) 占位符。"""
    return bool(s) and re.search(r'[一-鿿]', str(s)) and '无中文名' not in str(s)


# 中文名回填：库中存在「中文名实际为英文」的物质（多为风味数据库CSV来源），
# 按 CAS 号（优先）或规范化英文名匹配权威中文名，将原英文 cn 替换为中文。
_CN_BY_CAS = {
    '96-04-8': '2,3-庚二酮',                 # Acetyl valeryl = 2,3-heptanedione
    '156002-64-1': '杂环化合物',              # Heterocycles（泛称，无更具体中文）
    '2381-87-5': '脱水甲羟戊酸内酯',           # Dehydromevalonic lactone
    '4906-24-5': '2-乙酰氧基-3-丁酮',          # 2-acetoxy-3-butanone
    '394-32-1': '5-氟-2-羟基苯乙酮',           # 5-fluoro-2-hydroxy acetophenone
    '3796-70-1': '香叶基丙酮',                # Geranylacetone
    '461-55-2': '丁酸酯',                    # Butanoate
    '84-69-5': '邻苯二甲酸二异丁酯',           # Diisobutyl phthalate
    '143-20-4': '十六酸酯',                  # Hexadecanoate（棕榈酸酯）
    '5711-69-3': '异丁酸酯',                  # Isobutyrate
    '600-18-0': '2-氧代丁酸',                # 2-oxobutyric acid
    '56-84-8': '天冬氨酸',                   # Aspartic acid
    '77-92-9': '柠檬酸',                     # Citric acid
    '110-17-8': '富马酸',                    # Fumaric acid
    '56-86-0': '谷氨酸',                     # Glutamic acid
    '6915-15-7': '苹果酸',                   # Malic acid
    '144-62-7': '草酸',                      # Oxalic acid
    '127-17-3': '丙酮酸',                    # Pyruvic acid
    '110-15-6': '琥珀酸',                    # Succinic acid
    '87-69-4': '酒石酸',                     # Tartaric acid
    '57-55-6': '1,2-丙二醇',                 # 1,2-propanediol
    '513-85-9': '2,3-丁二醇',                # 2,3-butanediol
    '28343-22-8': '4-乙烯基丁香酚',           # 4-vinylsyringol
    '764-40-9': '2,4-戊二烯醛',              # 2,4-pentadienal
    '57568-60-2': '2-苯基-3-(2-呋喃基)丙烯醛',  # 2-Phenyl-3-(2-furyl)prop-2-enal
    '65505-17-1': '2-甲基-3-(甲基二硫基)呋喃',
    '614-18-6': '烟酸乙酯',
    '554-14-3': '2-甲基噻吩',
    '23074-10-4': '5-乙基糠醛',
    '21835-01-8': '3-乙基-1,2-环戊二酮',
    '765-70-8': '3-甲基-1,2-环戊二酮',
    '112-60-7': '四甘醇',
    '13529-27-6': '2-糠醛二乙缩醛',
    '21834-92-4': '5-甲基-2-苯基-2-己烯醛',
    '1741-41-9': '异丁醛二乙缩醛',
    '55088-52-3': 'α-乙叉基苯乙醛',
}
# 无 CAS 的物质按规范化英文名匹配（名称键已去除空格/连字符/大小写）
_CN_BY_EN = {
    'diepicedrene': '二表雪松烯',
    '2hydroxybutyrolactone': '2-羟基-γ-丁内酯',
    'heptalactone': 'γ-庚内酯',
    '3acetyl1pyrroline': '3-乙酰基-1-吡咯啉',
    'oxylene': '邻二甲苯',
    '56dihydro6pentyl2hpyran2one': '5,6-二氢-6-戊基-2H-吡喃-2-酮',
    'transcinnamicacid': '反式肉桂酸',
    '16octadien3ol': '1,6-辛二烯-3-醇',
    '2vinylethanol': '3-丁烯-1-醇',
    '7methanoazulen6ol': '7-甲桥薁-6-醇',
    'lmenthol': 'L-薄荷醇',
    'hept2enal': '2-庚烯醛',
    'non2enal': '2-壬烯醛',
    'oct2enal': '2-辛烯醛',
    'trans45epoxy': '反式-4,5-环氧-(E)-2-癸烯醛',
    'undecalactone': 'γ-十一内酯',
    '2butenoicacidethylester': '2-丁烯酸乙酯',
    'npropylbenzamide': 'N-正丙基苯甲酰胺',
    '3propanol': '正丙醇',
    '3hydroxy2methyl4pyranone': '3-羟基-2-甲基-4-吡喃酮',
    '3hydroxy2methyl4hpyran4one': '3-羟基-2-甲基-4H-吡喃-4-酮',
    '5hydroxymethylfurfural': '5-羟甲基糠醛',
    'methylpentadecaneether': '十五烷基甲基醚',
    'tetraethyleneglycol': '四甘醇',
    'methylbenzeneethanol': 'α-甲基苯乙醇',
    'methylbenzenemethanol': 'α-甲基苯甲醇',
    '2furaldehydediethylacetal': '2-糠醛二乙缩醛',
    '5methyl2phenyl2hexenal': '5-甲基-2-苯基-2-己烯醛',
    'benzeneacetaldehyde': '苯乙醛',
    'isobutyraldehydediethylacetal': '异丁醛二乙缩醛',
    'ethylidenebenzeneacetaldehyde': 'α-乙叉基苯乙醛',
    'methyl2methyl3furyldisulfide': '甲基-(2-甲基-3-呋喃基)二硫醚',
    'butanoicacid': '丁酸',
    'hexanoicacid': '己酸',
}
def _norm_key(s):
    return re.sub(r'[^a-z0-9]', '', (s or '').lower())
def _cn_from_cas():
    """将 cn 仍为英文的物质按 CAS / 英文名 回填为中文名。"""
    n = 0
    for c in COMPOUNDS:
        cn = c.get('cn') or ''
        if re.search(r'[一-鿿]', cn):
            continue  # 已有中文名，跳过
        cas = (c.get('cas') or '').strip()
        new = _CN_BY_CAS.get(cas) or _CN_BY_EN.get(_norm_key(c.get('en')))
        if new:
            c['cn'] = new
            n += 1
    if n:
        print('[info] 英文中文名回填(按CAS/名称): %d 条已转为中文' % n)


# ---------------------------------------------------------------------------
# 酯类 / 内酯类 中文名按 CAS 号修正
# 依据：化源网 chemsrc.com（https://www.chemsrc.com/）按 CAS 号查询得到的权威中文名，
#       并结合英文名(IUPAC/通用名)核定。
#       俗名差异（桃醛/γ-十一内酯、威士忌内酯、芥子油类、椰子醛、尼泊金酯等）
#       一律保留库中香料行业惯用名，不覆盖。
# ---------------------------------------------------------------------------
_ESTER_CN_BY_CAS = {
    # A. 醇基/酸基错配（库中把醇基或酸基张冠李戴）
    '623-42-7': '丁酸甲酯',      # Methyl butanoate —— 库中误作「丁酸乙酯」
    '105-79-3': '己酸异丁酯',    # 2-methylpropyl hexanoate —— 库中误作「己酸丙酯」
    '122-70-3': '丙酸苯乙酯',    # 2-phenylethyl propionate —— 库中误作「丙酸乙酯」

    # B. 名称残缺：仅剩「酸乙酯 / 酸甲酯 / 酸丙酯 / 乙基 / 甲基」等碎片
    '105-54-4': '丁酸乙酯',
    '3289-28-9': '环己烷甲酸乙酯',
    '110-38-3': '癸酸乙酯',
    '106-33-2': '月桂酸乙酯',
    '97-64-3': '乳酸乙酯',
    '106-32-1': '辛酸乙酯',
    '101-97-3': '苯乙酸乙酯',
    '105-37-3': '丙酸乙酯',
    '118-61-6': '水杨酸乙酯',
    '627-90-7': '十一烷酸乙酯',
    '1070-34-4': '琥珀酸单乙酯',
    '628-97-7': '棕榈酸乙酯',
    '52089-54-0': '2-羟基丁酸乙酯',
    '5405-41-4': '3-羟基丁酸乙酯',
    '999-10-0': '4-羟基丁酸乙酯',
    '638-10-8': '3-甲基-2-丁烯酸乙酯',
    '25415-67-2': '4-甲基戊酸乙酯',
    '103-36-6': '肉桂酸乙酯',
    '124-06-1': '十四酸乙酯',
    '617-05-0': '香草酸乙酯',
    '544-35-4': '亚油酸乙酯',
    '111-61-5': '硬脂酸乙酯',
    '111-62-6': '油酸乙酯',
    '614-99-3': '2-糠酸乙酯',
    '62696-37-1': '2-羟基-2-甲基-4-戊烯酸乙酯',
    '2441-06-7': '2-羟基-3-甲基丁酸乙酯',
    '10348-47-7': '2-羟基-4-甲基戊酸乙酯',
    '52089-55-1': '2-羟基己酸乙酯',
    '2021-28-5': '3-苯丙酸乙酯',
    '94-08-6': '对甲基苯甲酸乙酯',
    '14010-23-2': '十七烷酸乙酯',
    '41114-00-5': '十五酸乙酯',
    '93-58-3': '苯甲酸甲酯',
    '111-82-0': '月桂酸甲酯',
    '119-36-8': '水杨酸甲酯',
    '1534-08-3': '硫代乙酸-S-甲酯',
    '547-63-7': '异丁酸甲酯',
    '111-11-5': '辛酸甲酯',
    '24851-98-7': '二氢茉莉酮酸甲酯',
    '112-62-9': '油酸甲酯',
    '105-66-8': '丁酸丙酯',
    '626-77-7': '己酸丙酯',
    '16630-55-0': '乙酸-3-(甲硫基)丙酯',
    '141-04-8': '己二酸二异丁酯',
    '105-53-3': '丙二酸二乙酯',
    '2050-23-9': '辛二酸二乙酯',
    '120-47-8': '对羟基苯甲酸乙酯',
    '103-48-0': '异丁酸-2-苯乙酯',
    '19329-89-6': '乳酸异戊酯',

    # C. 两名拼接 或 Excel 断词残留的多余空格
    '108-64-5': '3-甲基丁酸乙酯',          # 原「甲基丁酸乙酯异戊酸乙酯」
    '97-62-1': '2-甲基丙酸乙酯',           # 原「甲基丙酸乙酯异丁酸乙酯」
    '15111-56-5': '3-环己烯-1-羧酸乙酯',
    '763-69-9': '3-乙氧基丙酸乙酯',
    '27829-72-7': '(E)-2-己烯酸乙酯',
    '26553-46-8': '(E)-3-己烯酸乙酯',
    '103-11-7': '丙烯酸-2-乙基己酯',
    '111-12-6': '2-辛炔酸甲酯',
    '137-05-3': '2-氰基丙烯酸甲酯',
    '7785-64-0': '(Z)-2-甲基-2-丁烯酸丁酯',  # 原 OCR 误「(2)-2-  甲基-2-丁酸丁酯」
    '2983-37-1': '2-乙基己酸乙酯',
    '39255-32-8': '2-甲基戊酸乙酯',
    '5870-68-8': '3-甲基戊酸乙酯',
    '78989-37-4': '(E)-4-辛烯酸乙酯',
    '10032-15-2': '2-甲基丁酸己酯',
    '27625-35-0': '2-甲基丁酸异戊酯',
    '2396-80-7': '5-己烯酸甲酯',
    '76238-22-7': '乙酸-(Z)-6-壬烯酯',
    '2442-10-6': '1-辛烯-3-醇乙酸酯',
    '68039-26-9': '2-甲基丁酸戊酯',
    '7699-00-5': 'D-乳酸乙酯',

    # D. 内酯缺 γ/δ 位标记，或含乱码字符
    '705-86-2': 'δ-癸内酯',
    '2305-05-7': 'γ-十二内酯',     # 与 713-95-1 的 δ-十二内酯 同名，须区分
    '713-95-1': 'δ-十二内酯',
    '695-06-7': 'γ-己内酯',        # 原「x己内酯」
    '698-76-0': 'δ-辛内酯',
    '3301-94-8': 'δ-壬内酯',
    '502-26-1': 'γ-十八内酯',      # 原「γ+八内酯」
    '1227-51-6': 'δ-十八内酯',
    '2721-22-4': 'δ-十四内酯',
    '710-04-3': 'δ-十一内酯',
    '823-22-3': 'δ-己内酯',
    '730-46-1': 'γ-十六内酯',      # 原 OCR 误「γ+六内酯」
    '2721-23-5': 'γ-十四内酯',     # 原 OCR 误「γ+四内酯」

    # E. 取代基缺位号 / 命名不规范
    '7452-79-1': '2-甲基丁酸乙酯',
    '644-49-5': '2-甲基丙酸丙酯',
    '53399-81-8': '2-甲基-4-戊烯酸乙酯',   # 与 4911-54-0 同名，须区分
    '4911-54-0': '4-甲基-4-戊烯酸乙酯',
    '10307-60-5': '(S)-2-甲基丁酸甲酯',
    '2305-25-1': '3-羟基己酸乙酯',
    '112-39-0': '棕榈酸甲酯',
    '106-70-7': '己酸甲酯',
    '1731-84-6': '壬酸甲酯',
    '84-66-2': '邻苯二甲酸二乙酯',        # 原「增塑剂 DEP」
    '688-84-6': '甲基丙烯酸-2-乙基己酯',
    '120-51-4': '苯甲酸苄酯',
    '626-82-4': '己酸丁酯',
    '14936-66-4': '乙酸-2-壬酯',
    '4192-77-2': '(E)-肉桂酸乙酯',     # 与 103-36-6 的肉桂酸乙酯 区分立体构型
    '1708-82-3': '(Z)-3-己烯醇乙酸酯',
    '21040-45-9': '(E)-乙酸肉桂酯',       # 与 103-54-8 同名，须区分
}

# 无 CAS 号、但中文名残缺的条目：按规范化英文名补回中文名
_ESTER_CN_BY_ENKEY = {
    'ethyl9octadecenoate': '油酸乙酯',
    'ethyloctadec9enoate': '油酸乙酯',
    'ethyloctadeca912dienoate': '亚油酸乙酯',
    'ethyl912octadecadienoate': '亚油酸乙酯',
    'ethyl91215octadecatrienoate': '亚麻酸乙酯',
    'ethyl3ethoxypropanoate': '3-乙氧基丙酸乙酯',
    'ethyl3phenylpropionate': '3-苯丙酸乙酯',
    'ethyl3phenylpropanoate': '3-苯丙酸乙酯',
    'ethylnicotinoate': '烟酸乙酯',
    'ethyl2hydroxypropanoate': '乳酸乙酯',
    '2hydroxyethylpropionate': '2-羟基乙基丙酸酯',
    'di2methylbutanedioate': '2-甲基丁二酸二乙酯',
    'smethyl3methylbutanethioate': 'S-甲基-3-甲基丁硫代酸酯',
}

# 名称残缺判定：只剩酸根或烷基碎片，无法辨识具体物质
_FRAG_CN = re.compile(
    r'^(?:酸[甲乙丙丁戊]?酯|甲基|乙基|丙基|丁基|甲基丙基|甲基乙基|乙基甲基|甲基甲基'
    r'|[甲乙丙]酯|己酸|丁酸)$')
_DUP_ESTER_CN = re.compile(r'酸[甲乙丙丁戊]?酯.*酸[甲乙丙丁戊]?酯')
_SPACES_CN = re.compile(r'\S\s{2,}\S')
_GARBAGE_CN = re.compile(r'(^[xX](?=[一-鿿])|[γΓδΔ]\+)')


def _fix_ester_lactone_cn():
    """酯类/内酯类中文名按 CAS 号修正（数据源 chemsrc.com）。
    仅对酯类/内酯类生效；未命中映射表的条目保持原名不动。"""
    n_cas = n_en = 0
    for c in COMPOUNDS:
        cat = (c.get('cat') or '')
        if not (cat.startswith('酯类') or cat.startswith('内酯')):
            continue
        old = (c.get('cn') or '').strip()
        new = _ESTER_CN_BY_CAS.get((c.get('cas') or '').strip())
        if new and new != old:
            c['cn'] = new
            n_cas += 1
            continue
        # 无 CAS 且中文名残缺 → 按英文名补回
        if not (c.get('cas') or '').strip() and (
                _FRAG_CN.match(old) or _DUP_ESTER_CN.search(old)
                or _SPACES_CN.search(old) or _GARBAGE_CN.search(old)):
            cn2 = _ESTER_CN_BY_ENKEY.get(_norm_key(c.get('en')))
            if cn2 and cn2 != old:
                c['cn'] = cn2
                n_en += 1
    if n_cas or n_en:
        print('[info] 酯类/内酯类中文名修正(按CAS): %d 条；按英文名补全(无CAS): %d 条'
              % (n_cas, n_en))


def _cas_valid(cas):
    """CAS 号校验位验证（如 106-33-0 校验位应为 2，属无效号）。"""
    s = re.sub(r'[^0-9]', '', (cas or ''))
    if len(s) < 3:
        return False
    body, chk = s[:-1], int(s[-1])
    return sum(int(d) * (i + 1) for i, d in enumerate(reversed(body))) % 10 == chk


def _dedupe_ester_lactone():
    """酯类/内酯类：中文名修正后暴露出的重复条目合并。
    判定同一物质：中文名完全相同。
    保留优先级：CAS 有效 > 有阈值 > 有气味描述 > 同义名数量多；
    被删条目的英文名并入保留条目同义名，缺失的阈值/介质补入保留条目。"""
    sel = [c for c in COMPOUNDS if (c.get('cat') or '').startswith('酯类')
           or (c.get('cat') or '').startswith('内酯')]
    if not sel:
        return

    def _rank(c):
        cas = (c.get('cas') or '').strip()
        return (1 if (cas and _cas_valid(cas)) else 0,
                1 if (c.get('thr_num') or 0) > 0 else 0,
                1 if (c.get('odor') or '').strip() else 0,
                len(c.get('syn') or []))

    def _can_merge(a, b):
        """两者 CAS 都有效且不同 → 视为不同物质，不合并。"""
        ca, cb = (a.get('cas') or '').strip(), (b.get('cas') or '').strip()
        va, vb = bool(ca) and _cas_valid(ca), bool(cb) and _cas_valid(cb)
        if va and vb:
            return ca == cb
        return True   # 至少一个 CAS 缺失/无效 → 按中文名合并

    groups = {}
    for c in sel:
        groups.setdefault((c.get('cn') or '').strip(), []).append(c)

    drop, merged = [], 0
    for cn, items in groups.items():
        if len(items) < 2:
            continue
        items_sorted = sorted(items, key=_rank, reverse=True)
        keep = items_sorted[0]
        for c in items_sorted[1:]:
            if not _can_merge(keep, c):
                continue
            # 英文名并入同义名
            en = (c.get('en') or '').strip()
            if en:
                syns = keep.setdefault('syn', []) or []
                if en.lower() not in [s.lower() for s in syns]:
                    syns.append(en)
            # 阈值补入
            if not (keep.get('thr_num') or 0) and (c.get('thr_num') or 0):
                keep['thr_num'] = c['thr_num']
                keep['thr'] = c.get('thr')
                keep['med'] = c.get('med')
                keep['thr_raw'] = c.get('thr_raw')
            # 气味描述补入
            if not (keep.get('odor') or '').strip() and (c.get('odor') or '').strip():
                keep['odor'] = c['odor']
            drop.append(c)
            merged += 1
    if drop:
        ids = {id(c) for c in drop}
        COMPOUNDS[:] = [c for c in COMPOUNDS if id(c) not in ids]
        print('[info] 酯类/内酯类重复条目合并: 删除 %d 条（英文名已并入保留条目同义名）' % merged)


def _fix_frag_cn():
    """全库扫描：中文名仅剩烷基/酸根碎片（无法辨识具体物质）的条目，
    按 CAS/英文名 回填为正确中文名（主要来源：化源网 chemsrc.com + 英文 IUPAC 名核定）。"""
    FRAG = re.compile(
        r'^(?:酸[甲乙丙丁戊]?酯|甲基|乙基|丙基|丁基|甲基丙基|甲基乙基|乙基甲基|甲基甲基'
        r'|[甲乙丙]酯|己酸|丁酸)$')
    n = 0
    for c in COMPOUNDS:
        old = (c.get('cn') or '').strip()
        if not FRAG.match(old):
            continue
        cas = (c.get('cas') or '').strip()
        new = _CN_BY_CAS.get(cas) or _CN_BY_EN.get(_norm_key(c.get('en')))
        if new and new != old:
            c['cn'] = new
            n += 1
    if n:
        print('[info] 残缺中文名按 CAS/EN 回填: %d 条' % n)



def _drop_no_cn():
    """删除无中文名的物质（cn 为 (无中文名)/英文碎片/空），仅从运行库剔除，不动源文件。
    发酵食品文献库(src='发酵食品文献库')的待核条目保留（用户明确要加入，有文献支撑）。"""
    before = len(COMPOUNDS)
    kept = []
    for c in COMPOUNDS:
        if _valid_cn(c.get('cn')):
            kept.append(c)
        elif c.get('_csv'):
            # flavor_database.csv 条目全量保留：该文件有 69 行的 name_cn 列填的是英文名
            #（如 Hexyl acetate、di-epi-cedrene），属文件本身写法，用户要求不去重全部加入。
            kept.append(c)
        elif c.get('src') in ('发酵食品文献库', '发酵食品文献库(1)') and c.get('_ferment_pending'):
            kept.append(c)   # 发酵食品待核物质保留
    dropped = before - len(kept)
    if dropped:
        COMPOUNDS[:] = kept
        print('[info] 已删除无中文名物质 %d 种（剩余 %d 种）' % (dropped, len(kept)))


def _drop_unknown():
    """删除未知化合物：无法确认身份的物质，包括
      1) 标记为待核(_ferment_pending)的条目（"待核化合物系统命名见英文"占位中文名）；
      2) 英文名以'-'开头的截断脏名（如 '-β-Ionone'、'1- propane'，Excel 解析残留）。
    仅从运行库剔除，不动源 json 文件。"""
    before = len(COMPOUNDS)
    kept = []
    for c in COMPOUNDS:
        en = (c.get('en') or '').strip()
        if c.get('_ferment_pending') or en.startswith('-'):
            continue
        kept.append(c)
    dropped = before - len(kept)
    if dropped:
        COMPOUNDS[:] = kept
        print('[info] 已删除未知化合物(待核/脏数据) %d 种（剩余 %d 种）' % (dropped, len(kept)))


def _expand_categories():
    """细分类别扩展：将库中「其他」或被粗分的物质按命名特征重新归类，
    使分类体系覆盖：含硫化合物、吡嗪类、噻唑类/噻唑啉类、呋喃/呋喃酮类、
    含氮杂环/其他含氮化合物、内酯类、萜烯类与含氧萜类。"""
    # 硫醇/硫醚/硫代（注意：thiazole 含 'thia'，但会在前面被噻唑规则分流）
    SULFUR = ('sulf', 'thiol', 'mercapt', 'disulfide', 'trisulfide', '硫',
              'furfurylthiol', 'furanthiol')
    # 含氮杂环（吡嗪/噻唑已单独识别）
    N_HETERO = ('pyridine', 'pyrrole', 'pyrrolidine', 'pyrimidine', 'indole',
                'quinoline', 'oxazole', 'isoxazole', 'pyrazole', 'imidazole',
                'azole', 'morpholine', 'piperidine', 'piperazine',
                '吡啶', '吡咯', '吲哚', '喹啉', '恶唑', '异恶唑', '吡唑',
                '咪唑', '氮杂', '吗啉', '哌啶', '哌嗪')
    n_changed = 0
    for c in COMPOUNDS:
        if c.get('_csv'):
            continue          # flavor_database.csv 条目：分类完全采用文件给定值，不按命名重分类
        en = (c.get('en') or '').lower()
        cn = (c.get('cn') or '')
        cur = (c.get('cat') or '其他').strip()
        new = None
        # 只处理可以被细化的当前类别
        if cur not in ('其他', '酯类', '酮类', '醇类', '醛类', '酸类', '萜烯类'):
            continue
        if 'pyrazine' in en or '吡嗪' in cn:
            new = '吡嗪类'
        elif 'thiazole' in en or 'thiazoline' in en or '噻唑' in cn:
            new = '噻唑类 / 噻唑啉类'
        elif any(h in en for h in SULFUR) or '硫' in cn:
            new = '含硫化合物'
        elif any(h in en for h in ('furan', 'furaneol', 'furanone', 'furfural', 'furfuryl')) or '呋喃' in cn:
            new = '呋喃/呋喃酮类'
        elif any(h in en for h in N_HETERO):
            new = '含氮杂环/其他含氮化合物'
        elif 'lactone' in en or '内酯' in cn:
            new = '内酯类'
        elif any(h in en for h in TERPENE_HINTS):
            new = '萜烯类与含氧萜类'
        if new and new != cur:
            c['cat'] = new
            n_changed += 1
    if n_changed:
        print('[info] 细分类别扩展: %d 种物质重新归类' % n_changed)


def _normalize_categories():
    """统一类别名称：把旧的/简写类别名合并为用户指定的新名称。"""
    MAP = {
        '萜烯类': '萜烯类与含氧萜类',
    }
    n = 0
    for c in COMPOUNDS:
        cat = (c.get('cat') or '其他').strip()
        if cat in MAP:
            c['cat'] = MAP[cat]
            n += 1
    if n:
        print('[info] 类别名称统一: %d 种物质从「萜烯类」合并为「萜烯类与含氧萜类」' % n)


# 数据库来源 -> 可读"来源"标注（依据上文给出的数据库文件命名）
_SRC_MAP = {
    'FTDB v1.5.0': 'FTDB 风味数据库',
    '发酵食品文献库': '发酵食品挥发性物质文献数据库',
    '发酵食品文献库(1)': '发酵食品挥发性物质文献数据库(修订版)',
    'Van Gemert 2011': 'Van Gemert 2011 风味阈值汇编',
    'HTML表(风味物质完整版)': '挥发性风味物质完整版阈值CAS对照表',
    '图片': '图片标注',
    '自定义': '自定义录入',
    '—': '—',
    '': '—',
}


def dedupe_phrases(s):
    """按 '、；;，' 拆分短语并去重（保留首次出现顺序），用于「主要来源/来源」字段的
    重复字符/短语自动删除。例如 '熟肉、高汤；熟肉' -> '熟肉、高汤'。
    不按空白拆分，以免破坏含空格的数据来源名（如 'Van Gemert 2011 风味阈值汇编'）。"""
    if not s:
        return ''
    seen, out = set(), []
    for p in re.split(r'[、；;，]', s):
        p = p.strip()
        if not p:
            continue
        if p not in seen:
            seen.add(p)
            out.append(p)
    return '、'.join(out)


# ---- 来源按食品类别分组（一类一行） ----
FOOD_CAT_ORDER = ['肉类', '乳制品', '茶类', '柑橘类', '酒类', '调味品', '其他']


def food_cat(p):
    """将单个来源产品短语归类到食品类别（肉类/乳制品/茶类/柑橘类/酒类/调味品/其他）。"""
    x = p
    if any(k in x for k in ['牛乳', '黄油', '稀奶油', '奶油', '发酵乳', '奶粉', '乳',
                            '奶酪', '干酪', '热处理乳', 'UHT', '低脂乳', '发酵黄油']):
        return '乳制品'
    if any(k in x for k in ['橙', '柚', '柑', '柠檬', '葡萄柚', '宽皮柑橘', '柑橘']):
        return '柑橘类'
    if any(k in x for k in ['茶', '茶汤', '叶（']):
        return '茶类'
    if any(k in x for k in ['白酒', '米酒', '黄酒', '陈酿', '酒']):
        return '酒类'
    if any(k in x for k in ['酱油', '蚝油', '食醋', '醋', '芝麻油']):
        return '调味品'
    if any(k in x for k in ['肉类', '牛肉', '鸡肉', '猪肉', '羊肉', '火腿', '高汤', '肉汤',
                            '炖', '和牛', '鱼', '海鲜', '油脂', '脂质', '肉',
                            '烤香', '食品挥发', '发酵/烤']):
        return '肉类'
    return '其他'


def _split_protect(s):
    """按 '、，,/／；;' 拆分，但保护括号内的内容（如 '白酒（浓香型、陈酿）' 不被拆破）。"""
    if not s:
        return []
    parts, buf, depth = [], '', 0
    for ch in s:
        if ch in '（(':
            depth += 1; buf += ch
        elif ch in '）)':
            depth -= 1; buf += ch
        elif ch in '、，,/／；;' and depth == 0:
            if buf.strip():
                parts.append(buf.strip())
            buf = ''
        else:
            buf += ch
    if buf.strip():
        parts.append(buf.strip())
    return parts


def group_source_lines(phrases):
    """将一个物质的所有来源产品短语，按食品类别分组，返回 ['类别：p1、p2', ...]（一类一行）。"""
    buckets = {c: [] for c in FOOD_CAT_ORDER}
    for raw in phrases:
        for p in _split_protect(raw):
            c = food_cat(p)
            if c != '其他' and p not in buckets[c]:
                buckets[c].append(p)
    return [c + '：' + '、'.join(buckets[c]) for c in FOOD_CAT_ORDER if buckets[c]]


def build_source(dist_by_cat, products):
    """构建「来源」字段：把 HTML 分段(dist_by_cat)与发酵食品报道产品(products)中的
    每个产品短语，统一按 food_cat 重新归类，最终一类一行（\\n 分隔）；
    无产品列表时返回空串（交由调用方回退 DB 名）。无法归入食品类别的短语(其他)不显示。"""
    dist_by_cat = dist_by_cat or {}
    buckets = {c: [] for c in FOOD_CAT_ORDER}
    tokens = []
    for val in dist_by_cat.values():
        items = val if isinstance(val, list) else _split_protect(val)
        tokens.extend(items)
    if products:
        tokens.extend(_split_protect(products))
    for p in tokens:
        p = p.strip()
        if not p:
            continue
        c = food_cat(p)
        if c != '其他' and p not in buckets[c]:
            buckets[c].append(p)
    lines = [c + '：' + '、'.join(buckets[c]) for c in FOOD_CAT_ORDER if buckets.get(c)]
    return '\n'.join(lines)


def _dedupe_odor():
    """删除气味描述(odor)中的重复字眼，并处理粘连/异形分隔导致的叠词。

    做法：
    1) 由库中已有的独立风味词构建「词表」，并迭代剔除可被拆成≥2个已知词的粘连 token
       （如「青草香脂肪香」→ 青草香+脂肪香，「清新香气青香青草香」→ 清新香气+青香+青草香），
       化学类名（芳香酮、香料、香菜气味等）因无法拆成已知词而保持完整。
    2) 对每条气味描述做前向最长匹配分词（连字符/空格/冒号作分隔，粘连串被拆开），
       跨段保序去重；同时清除「其他特征」这类粘连填充词。
    3) 保留「风味类别: …、描述: …」的结构标签，仅删除跨标签重复的词。"""
    import re as _re
    SEP = r'[、,，/／；;－\-]'
    LABEL = _re.compile(r'(?:风味类别|描述)\s*[:：]')
    STRIP = '、，,/／；;－- :：'

    # ---- 构建词表（迭代剔除可被拆分的粘连 token） ----
    _leaves = set()
    for c in COMPOUNDS:
        s = (c.get('note') or '').strip()
        if not s or s == '—':
            continue
        s = LABEL.sub('', s)
        for part in _re.split(SEP, s):
            t = part.strip().rstrip('-').strip()
            t = t.replace('其他特征', '')
            if t:
                _leaves.add(t)

    def _fmm(text, dic):
        i, n = 0, len(text); out = []
        while i < n:
            m = None
            for w in sorted(dic, key=len, reverse=True):
                if text.startswith(w, i):
                    m = w; break
            if not m:
                return None
            out.append(m); i += len(m)
        return out

    changed = True
    while changed:
        changed = False
        for w in list(_leaves):
            if len(w) <= 2:
                continue
            seg = _fmm(w, _leaves - {w})
            if seg and len(seg) >= 2:
                _leaves.discard(w)
                for p in seg:
                    _leaves.add(p)
                changed = True
    _vl = sorted(_leaves, key=len, reverse=True)
    _sep_set = set(STRIP)

    def _segment(text):
        i, n = 0, len(text); out = []
        while i < n:
            if text[i] in _sep_set:
                i += 1; continue
            if text.startswith('其他特征', i):   # 粘连填充词，直接跳过
                i += 4; continue
            m = None
            for w in _vl:
                if text.startswith(w, i):
                    m = w; break
            if m:
                out.append(m); i += len(m)
            else:
                j = i
                while j < n and text[j] not in _sep_set:
                    j += 1
                out.append(text[i:j]); i = j
        return out

    for c in COMPOUNDS:
        raw = (c.get('note') or '').strip()
        if not raw or raw == '—':
            continue
        # 结构化：分别处理 风味类别 / 描述 两段（段内分词、跨段去重）
        if '风味类别' in raw:
            idx = raw.find('描述') if '描述' in raw else -1
            if idx >= 0:
                cat_seg = LABEL.sub('', raw[:idx]).strip(STRIP)
                desc_seg = LABEL.sub('', raw[idx + len('描述'):]).strip(STRIP)
            else:
                cat_seg = LABEL.sub('', raw).strip(STRIP); desc_seg = ''
            cat_words = _segment(cat_seg)
            seen = set(cat_words)
            desc_words = [w for w in _segment(desc_seg)
                          if w not in seen and not seen.add(w)]
            cat_str = '、'.join(cat_words)
            c['note'] = ('风味类别: %s、描述: %s' % (cat_str, '、'.join(desc_words))) \
                if desc_words else ('风味类别: %s' % cat_str)
            continue
        # 非结构化：全局分词 + 保序去重
        words = _segment(raw)
        c['note'] = '、'.join(dict.fromkeys(words))


def _finalize_columns():
    """将备注拆分为「气味描述」与「来源」两栏，CAS 置于中文名右侧，统一字段顺序。
    字段顺序：英文名 / 中文名 / CAS / 类别 / 阈值 / 介质 / 气味描述 / 来源 / (内部: src,syn)。
    气味描述 = 原 note 中的气味词；来源 = 发酵食品文献库物质显示"报道产品"，其余显示数据库来源名。"""
    for c in COMPOUNDS:
        odor = (c.get('note') or '').strip()
        if not odor or odor == '—':
            odor = c.get('note') or '—'
        src_raw = (c.get('src') or '—').strip()
        # 来源：按食品类别分组，一类一行（肉类/乳制品/茶类/柑橘类/酒类/调味品）
        # 优先 HTML 分段(dist_by_cat，权威) + 发酵食品报道产品(products，按类别并入)；
        # 无产品列表时回退为数据库来源名。
        prod = (c.get('products') or '').strip()
        dist_by_cat = c.get('dist_by_cat') or {}
        grouped = build_source(dist_by_cat, prod)
        if grouped:
            source = grouped
        elif src_raw.startswith('发酵食品文献库') and prod:
            source = prod
        else:
            source = _SRC_MAP.get(src_raw, src_raw)
        # 重组字段顺序（保留内部字段 src/syn/cas 等）
        new = {
            'en': c.get('en', ''),
            'cn': c.get('cn', ''),
            'cas': (c.get('cas') or '').strip(),
            'cat': c.get('cat', '其他'),
            'thr': c.get('thr', '—'),
            'med': c.get('med', '—'),
            'thr_num': c.get('thr_num'),                  # μg/L 数值（OAV/ROAV 计算用）
            'thr_raw': (c.get('thr_raw') or '').strip(),  # 折算前的原始阈值文本
            '_csv': c.get('_csv', False),                 # 标记 flavor_database.csv 条目
            'odor': odor,          # 气味描述（原备注）
            'source': source,      # 来源（按食品类别分组，一类一行）
            'products': prod,      # 报道产品（内部保留）
            'dist': (c.get('dist') or '').strip(),      # 主要来源并集（内部保留）
            'dist_by_cat': dist_by_cat,                 # 主要来源分段（内部保留）
            'src': src_raw,        # 内部保留
            'syn': c.get('syn') or [],
        }
        c.clear()
        c.update(new)



def _dedupe_pyrazine_cas(cat='吡嗪类'):
    """删除「吡嗪类」中 CAS 号重复的条目，同一 CAS 仅保留一条。

    用户要求：重复物质优先采用 flavor_database.csv 的数据。保留规则：
      1) 优先保留来源为『风味数据库CSV』的条目；
      2) 若同组多条均来自该 CSV（或均不是），按数据完整度打分取最高
         —— 阈值(8) > 气味描述(6) > 来源/产品(3) > 介质(2) > 中文名无括号(2)
            > 英文名含位号(2) > 同义名(1)；
      3) 仍并列则保留最先出现者（稳定）。
    被删条目的英文名/中文名（含括号内别名）并入保留条目的 syn，保证 GC-MS 报告仍能匹配；
    不并入被删条目的 syn 列表（可能含跨库污染的无关别名，会造成误匹配）。
    保留条目中为空('—'/空串)的字段用被删条目的值回填，避免既有信息丢失。"""
    def _ck(c):
        s = (c.get('cas') or '').strip().replace('-', '').lower()
        return s if s and s != '—' else ''

    def _score(c):
        s = 0
        if (c.get('src') or '') == '风味数据库CSV':
            s += 100                                  # 用户指定数据源优先
        if (c.get('thr') or '').strip() not in ('', '—'):
            s += 8
        if (c.get('med') or '').strip() not in ('', '—'):
            s += 2
        if (c.get('odor') or '').strip() not in ('', '—'):
            s += 6
        prod = (c.get('products') or '').strip()
        srcv = (c.get('source') or '').strip()
        if prod or (srcv and srcv != '—'):
            s += 3
        cn = c.get('cn') or ''
        if '（' not in cn and '(' not in cn:
            s += 2                                    # 规范中文名（无括号备注）
        if re.search(r'\d', c.get('en') or ''):
            s += 2                                    # 英文名带位号，更精确
        if c.get('syn'):
            s += 1
        return s

    groups = {}
    for i, c in enumerate(COMPOUNDS):
        if c.get('_csv'):
            continue          # flavor_database.csv 条目按用户要求不去重，原样保留
        if (c.get('cat') or '').strip() != cat:
            continue
        ck = _ck(c)
        if not ck:
            continue
        groups.setdefault(ck, []).append(i)

    drop, removed_names = set(), []
    for ck, idxs in groups.items():
        if len(idxs) < 2:
            continue
        keep = max(idxs, key=lambda i: (_score(COMPOUNDS[i]), -i))
        surv = COMPOUNDS[keep]
        syn = list(surv.get('syn') or [])
        for i in idxs:
            if i == keep:
                continue
            other = COMPOUNDS[i]
            # 被删条目的名称并入同义名，保证 GC-MS 报告仍可匹配。
            # 只取条目自身的 en/cn 及其括号内的别名（如 (pepper pyrazine)、（川芎嗪）），
            # 【不并入其 syn 列表】——该列表可能混入其它数据源的无关别名
            #（例如发酵食品文献库把 'diisobutyl ketone'、'2,6-dimethyl-4-heptanone'
            #  挂在 METHYLPYRAZINE 条目下），并入会让无关物质精确误匹配到本条目。
            names = []
            for nm in (other.get('en'), other.get('cn')):
                nm = (nm or '').strip()
                if nm:
                    names.append(nm)
                    for al in re.findall(r'[（(]([^）)]*)[）)]', nm):
                        al = al.strip()
                        if al:
                            names.append(al)
            for nm in names:
                if nm and nm not in syn:
                    syn.append(nm)
            # 空字段回填（仅当保留条目该字段为空时才用被删条目的值）
            for f in ('thr', 'med', 'odor'):
                cur = (surv.get(f) or '').strip()
                if cur in ('', '—'):
                    v = (other.get(f) or '').strip()
                    if v and v != '—':
                        surv[f] = v
            if not (surv.get('products') or '').strip():
                v = (other.get('products') or '').strip()
                if v:
                    surv['products'] = v
            removed_names.append((other.get('en') or '') + '(CAS ' + (other.get('cas') or '') + ')')
            drop.add(i)
        surv['syn'] = syn

    if drop:
        COMPOUNDS[:] = [c for i, c in enumerate(COMPOUNDS) if i not in drop]
        print('[info] %s CAS 去重: 删除重复条目 %d 条（名称已并入保留条目同义名）: %s'
              % (cat, len(drop), '、'.join(removed_names)))


def _clean_pyrazine_syns(cat='吡嗪类'):
    """清理「吡嗪类」条目的同义名，只保留与吡嗪相关的别名（含 razine / 嗪）。

    该类别的 syn 里混有两类脏数据：
      1) 跨库污染：其它数据源把无关物质的别名挂到本条目下
         （如发酵食品文献库把 'diisobutyl ketone'、'2,6-dimethyl-4-heptanone'
           挂在 2-甲基吡嗪下，会导致报告中的酮类被精确误匹配成吡嗪）；
      2) 名称拆分碎片：由英文名括号/逗号切出的取代基片段
         （'2-methylbutyl'、'3-(e'、'phenylthio'、's' 等）。
         这类短片段一旦进入检索索引，会让大量含该片段的无关物质被「包含匹配」误命中。
    保留规则：别名含 razine（覆盖 pyrazine / ligustrazine / trimethylpyrazine …）
    或含「嗪」（覆盖 川芎嗪 / 甲基吡嗪 …）即视为吡嗪相关。"""
    pat = re.compile(r'razine|嗪', re.I)
    n_drop = 0
    for c in COMPOUNDS:
        if c.get('_csv'):
            continue          # flavor_database.csv 条目按用户要求原样保留
        if (c.get('cat') or '').strip() != cat:
            continue
        syn = c.get('syn') or []
        keep = [s for s in syn if pat.search(s or '')]
        if len(keep) != len(syn):
            n_drop += len(syn) - len(keep)
            c['syn'] = keep
    if n_drop:
        print('[info] %s 同义名清洗: 移除无关/碎片别名 %d 个（仅保留含 razine/嗪 的别名）' % (cat, n_drop))


def _dedupe_cas_global():
    """全局 CAS 去重：相同 CAS 只保留一条。

    规则：
      1) 若重复组内存在 FTDB v1.5.0 来源的条目，先删除 FTDB 条目；
      2) 在剩余条目中按「来源明确程度 + 数据完整度」打分，保留最高者；
      3) 被删条目的自身名称（en/cn 及其括号内别名）并入保留条目 syn，
         保证 GC-MS 报告仍可匹配；不并入被删条目的完整 syn 列表，避免污染。
      4) 保留条目为空的字段（阈值/介质/气味/来源/产品）用被删条目的值回填。
    """
    FTDB_SRC = 'FTDB v1.5.0'

    def _ck(c):
        s = (c.get('cas') or '').strip().replace('-', '').lower()
        return s if s and s != '—' else ''

    def _is_placeholder(v):
        return (v or '').strip() in ('', '—', '-', '–', 'N/A', 'na', 'NULL')

    def _score(c):
        s = 0
        src = (c.get('src') or '').strip()
        # 来源明确程度
        if src and src != '—':
            s += 10
        if src == FTDB_SRC:
            s -= 50                     # FTDB 明确降级
        # 数据完整度
        if not _is_placeholder(c.get('cn')):
            s += 4
        if not _is_placeholder(c.get('en')):
            s += 2
        if not _is_placeholder(c.get('thr')):
            s += 3
        if not _is_placeholder(c.get('odor')):
            s += 2
        if not _is_placeholder(c.get('med')):
            s += 1
        if (c.get('products') or '').strip() or (c.get('source') or '').strip():
            s += 1
        if c.get('syn'):
            s += min(len(c['syn']), 3) * 0.5
        return s

    groups = {}
    for i, c in enumerate(COMPOUNDS):
        ck = _ck(c)
        if not ck:
            continue
        groups.setdefault(ck, []).append(i)

    drop, removed_names = set(), []
    for ck, idxs in groups.items():
        if len(idxs) < 2:
            continue
        # 阶段1：如果组内同时存在 FTDB 与非 FTDB，删除所有 FTDB
        non_ftdb = [i for i in idxs if (COMPOUNDS[i].get('src') or '').strip() != FTDB_SRC]
        if non_ftdb and len(non_ftdb) < len(idxs):
            for i in idxs:
                if i not in non_ftdb:
                    drop.add(i)
                    removed_names.append((COMPOUNDS[i].get('en') or '') + '(CAS ' + (COMPOUNDS[i].get('cas') or '') + ')')
            idxs = non_ftdb
        if len(idxs) < 2:
            continue
        # 阶段2：按评分保留最佳
        keep = max(idxs, key=lambda i: (_score(COMPOUNDS[i]), -i))
        surv = COMPOUNDS[keep]
        syn = list(surv.get('syn') or [])
        for i in idxs:
            if i == keep:
                continue
            other = COMPOUNDS[i]
            names = []
            for nm in (other.get('en'), other.get('cn')):
                nm = (nm or '').strip()
                if nm:
                    names.append(nm)
                    for al in re.findall(r'[（(]([^）)]*)[）)]', nm):
                        al = al.strip()
                        if al:
                            names.append(al)
            for nm in names:
                if nm and nm not in syn:
                    syn.append(nm)
            for f in ('thr', 'med', 'odor', 'src'):
                if _is_placeholder(surv.get(f)):
                    v = (other.get(f) or '').strip()
                    if not _is_placeholder(v):
                        surv[f] = v
            # 若保留条目缺少中文名，而被删条目有中文名，则用中文名回填
            surv_cn = surv.get('cn') or ''
            other_cn = (other.get('cn') or '').strip()
            if not re.search(r'[一-鿿]', surv_cn) and re.search(r'[一-鿿]', other_cn):
                surv['cn'] = other_cn
            if not (surv.get('products') or '').strip():
                v = (other.get('products') or '').strip()
                if v:
                    surv['products'] = v
            removed_names.append((other.get('en') or '') + '(CAS ' + (other.get('cas') or '') + ')')
            drop.add(i)
        surv['syn'] = syn

    if drop:
        COMPOUNDS[:] = [c for i, c in enumerate(COMPOUNDS) if i not in drop]
        print('[info] 全局 CAS 去重: 删除重复条目 %d 条（名称已并入保留条目同义名）: %s'
              % (len(drop), '、'.join(removed_names[:20]) + (' 等' if len(removed_names) > 20 else '')))


def _drop_artifacts():
    """删除分析前处理带入的溶剂残留、抗氧化剂(BHT)衍生类等非风味 artifact 物质。
    按 note 中的精确负面短语判定，仅从运行库剔除，不动源文件。"""
    rules = ['溶剂残留/前处理带入', '前处理带入', 'BHT 氧化产物',
             '可能来自抗氧化剂 BHT 衍生', '可能来自橡胶抗氧化剂迁移']
    before = len(COMPOUNDS)
    kept, removed = [], []
    for c in COMPOUNDS:
        note = (c.get('note') or '')
        if any(p in note for p in rules):
            removed.append(c.get('en'))
        else:
            kept.append(c)
    if removed:
        COMPOUNDS[:] = kept
        print('[info] 已删除溶剂残留/BHT衍生等 artifact 物质 %d 种: %s'
              % (len(removed), ', '.join(removed)))


def _load_curated_extras():
    """合并 curated_extras.json（来自图片表的物质清单）。
    - 已存在的条目（含归一化英文名匹配或 CAS 相同）：补充风味描述与阈值，不覆盖；
    - 未收录的物质：追加为新条目，分类按图标注，来源标 "图片" 便于追溯。"""
    p = os.path.join(HERE, 'curated_extras.json')
    if not os.path.exists(p):
        return
    try:
        raw = json.load(open(p, 'r', encoding='utf-8'))
    except Exception as e:
        print('[warn] 读取 curated_extras.json 失败:', e)
        return
    items = [x for x in raw.get('items', []) if not x.get('_skip') and x.get('en')]

    # 建立库索引（按归一化英文名/中文名/CAS 三键）
    def keyof(c):
        return normalize(c.get('en', '')) or normalize(c.get('cn', '')) or (c.get('cas') or '').lower()
    by_en, by_cas = {}, {}
    for i, c in enumerate(COMPOUNDS):
        k = normalize(c.get('en', ''))
        if k:
            by_en.setdefault(k, []).append(i)
        ck = (c.get('cas') or '').lower().replace('-', '').strip()
        if ck:
            by_cas.setdefault(ck, []).append(i)

    added, updated = 0, 0
    for x in items:
        en, cn = x.get('en', '').strip(), x.get('cn', '').strip()
        cas = (x.get('cas') or '').strip()
        thr = (x.get('thr') or '').strip()
        med = (x.get('med') or '').strip()
        cat = (x.get('cat') or '其他').strip()
        note = (x.get('note') or '').strip()
        nk = normalize(en)
        ck = cas.lower().replace('-', '').strip()

        # 查找已有条目
        hits = by_en.get(nk, [])
        if not hits and ck:
            hits = by_cas.get(ck, [])
        target = None
        if hits:
            # 优先选未标 src='图片' 的（即首次出现=精选或 FTDB）
            for i in hits:
                if COMPOUNDS[i].get('src') != '图片':
                    target = COMPOUNDS[i]; break
            if target is None:
                target = COMPOUNDS[hits[0]]

        # 把图片风味描述标准化成一句话
        if note.startswith('Odor:') or note.startswith('Odor描述'):
            flavor_desc = note
        elif '；' in note:
            flavor_desc = note
        else:
            flavor_desc = note

        if target:
            # 补充（不覆盖已有值）
            prev_note = target.get('note', '')
            tag = '图片风味描述: ' + flavor_desc
            if tag not in prev_note:
                target['note'] = (prev_note.rstrip('；') + '；' + tag).strip('；')
            # 阈值：若图片给出有效阈值，原 thr 进备注，th 替换为图片值
            if thr not in ('', '—'):
                cur = (target.get('thr') or '').strip()
                if cur and cur != thr:
                    target['note'] = (target['note'].rstrip('；') + '；原thr=' + cur).strip('；')
                target['thr'] = thr
            if med not in ('', '文献未列'):
                cur = (target.get('med') or '').strip()
                if cur in ('', '文献未列'):
                    target['med'] = med
            # CAS 兜底
            if cas and not target.get('cas'):
                target['cas'] = cas
            # 中文名兜底：若现有 cn 为空、纯数字或与图片 cn 冲突，以图片为准
            cur_cn = (target.get('cn') or '').strip()
            if cn and (not cur_cn or cur_cn.isdigit() or len(cur_cn) <= 1):
                target['cn'] = cn
            # 类别兜底：若现有 cat 为"其他"或缺失，采纳图片标注
            cur_cat = (target.get('cat') or '').strip()
            if cat and cat != '其他' and cur_cat in ('', '其他'):
                target['cat'] = cat
            target['_img_src'] = True
            updated += 1
        else:
            # 新增条目
            COMPOUNDS.append({
                'en': en, 'cn': cn, 'cat': cat, 'thr': thr if thr else '—',
                'med': med if med else '文献未列',
                'note': flavor_desc,
                'src': '图片', 'syn': [],
                'cas': cas,
            })
            added += 1
            if nk:
                by_en.setdefault(nk, []).append(len(COMPOUNDS) - 1)
            if ck:
                by_cas.setdefault(ck, []).append(len(COMPOUNDS) - 1)

    if added or updated:
        print('[info] curated_extras: 新增 %d, 补充 %d, 合计处理 %d' % (added, updated, len(items)))


def _load_volatile_html():
    """合并 volatile_flavor_html.json（解析自《挥发性风味物质完整版 阈值CAS对照表》HTML）。
    - 已存在条目（归一化英文名 / CAS 匹配）：补充缺失的 CAS、中文名、水相阈值、风味描述，不覆盖已有值；
    - 未收录物质：追加为新条目，src 标 'HTML表(风味物质完整版)' 便于追溯。"""
    p = os.path.join(HERE, 'volatile_flavor_html.json')
    if not os.path.exists(p):
        return
    try:
        raw = json.load(open(p, 'r', encoding='utf-8'))
    except Exception as e:
        print('[warn] 读取 volatile_flavor_html.json 失败:', e)
        return
    items = [x for x in raw.get('items', []) if x.get('en')]

    by_en, by_cas = {}, {}
    for i, c in enumerate(COMPOUNDS):
        k = normalize(c.get('en', ''))
        if k:
            by_en.setdefault(k, []).append(i)
        ck = (c.get('cas') or '').lower().replace('-', '').strip()
        if ck:
            by_cas.setdefault(ck, []).append(i)

    added, updated = 0, 0
    for x in items:
        en = (x.get('en') or '').strip()
        cn = (x.get('cn') or '').strip()
        cas = (x.get('cas') or '').strip()
        thr = (x.get('thr') or '').strip()
        med = (x.get('med') or '').strip()
        cat = (x.get('cat') or '其他').strip()
        note = (x.get('note') or '').strip()
        dist = (x.get('dist') or '').strip()
        nk = normalize(en)
        ck = cas.lower().replace('-', '').strip()

        hits = by_en.get(nk, [])
        if not hits and ck:
            hits = by_cas.get(ck, [])
        target = None
        if hits:
            for i in hits:
                if COMPOUNDS[i].get('src') != 'HTML表(风味物质完整版)':
                    target = COMPOUNDS[i]; break
            if target is None:
                target = COMPOUNDS[hits[0]]

        if target:
            # 风味描述：HTML 来源（多为文献实测）补充进 note，按句去重，避免标签/重复堆积
            if note and note != '—':
                prev = (target.get('note') or '').strip()
                merged = (note + '；' + prev) if prev not in ('', '—') else note
                parts = [p.strip() for p in re.split(r'[；;]', merged) if p.strip()]
                seen = set(); dedup = []
                for p in parts:
                    if p not in seen:
                        seen.add(p); dedup.append(p)
                target['note'] = '；'.join(dedup)
            # 阈值：水相优先；现有无阈值时补入
            if thr not in ('', '—'):
                cur = (target.get('thr') or '').strip()
                if cur in ('', '—'):
                    target['thr'] = thr
                    if med not in ('', '—'):
                        target['med'] = med
                elif med not in ('', '—') and (target.get('med') in ('', '—', '文献未列')):
                    target['med'] = med
            # CAS 兜底
            if cas and not target.get('cas'):
                target['cas'] = cas
            # 中文名兜底（现有无有效中文名时补；"(无中文名)"视为无效）
            cur_cn = (target.get('cn') or '').strip()
            def _valid_cn(s):
                return bool(s) and re.search(r'[一-鿿]', s) and '无中文名' not in s
            if cn and not _valid_cn(cur_cn):
                target['cn'] = cn
            # 主要来源(dist)补充：与已有 dist 取并集并按短语去重（重复字符自动删除）
            if dist and dist != '—':
                prev_dist = (target.get('dist') or '').strip()
                target['dist'] = dedupe_phrases((prev_dist + '；' + dist) if prev_dist else dist)
            # 主要来源分段(dist_by_cat)合并：同一物质跨食品段并入对应类别
            xdbc = x.get('dist_by_cat') or {}
            if xdbc:
                prev_dbc = target.get('dist_by_cat') or {}
                for k, v in xdbc.items():
                    lst = prev_dbc.setdefault(k, [])
                    for p in re.split(r'[、，,/／；;]', v):
                        p = p.strip()
                        if p and p not in lst:
                            lst.append(p)
                target['dist_by_cat'] = prev_dbc
            # 类别兜底：HTML 细分类别优先于「其他」；内酯类从「酯类」析出
            cur_cat = (target.get('cat') or '').strip()
            if cat and cat != '其他' and (cur_cat in ('', '其他') or (cur_cat == '酯类' and cat == '内酯类')):
                target['cat'] = cat
            updated += 1
        else:
            COMPOUNDS.append({
                'en': en, 'cn': cn, 'cat': cat,
                'thr': thr if thr else '—', 'med': med if med else '文献未列',
                'note': note if note else '—',
                'src': 'HTML表(风味物质完整版)', 'syn': x.get('syn') or [],
                'cas': cas, 'dist': dist if dist else '',
                'dist_by_cat': x.get('dist_by_cat') or {},
            })
            added += 1

    if added or updated:
        print('[info] volatile_flavor_html: 新增 %d, 补充 %d, 合计处理 %d' % (added, updated, len(items)))


def _load_fermented_food():
    """合并 fermented_food_db.json（发酵食品挥发性物质文献数据库 xlsx）。
    用户要求优先用该文件覆盖：匹配到的现有条目，强制以本文件的
    中文名(非待核)/类别/阈值/介质/风味描述覆盖；未收录的追加为 src='发酵食品文献库'。"""
    p = os.path.join(HERE, 'fermented_food_db.json')
    if not os.path.exists(p):
        return
    try:
        items = json.load(open(p, 'r', encoding='utf-8'))
    except Exception as e:
        print('[warn] 读取 fermented_food_db.json 失败:', e)
        return

    by_en, by_cas = {}, {}
    for i, c in enumerate(COMPOUNDS):
        k = normalize(c.get('en', ''))
        if k:
            by_en.setdefault(k, []).append(i)
        ck = (c.get('cas') or '').lower().replace('-', '').strip()
        if ck:
            by_cas.setdefault(ck, []).append(i)

    added, updated = 0, 0
    for x in items:
        en = (x.get('en') or '').strip()
        cn = (x.get('cn') or '').strip()
        cas = (x.get('cas') or '').strip()
        cat = (x.get('cat') or '其他').strip()
        thr = (x.get('thr') or '').strip()
        med = (x.get('med') or '').strip()
        note = (x.get('note') or '').strip()
        nk = normalize(en)
        ck = cas.lower().replace('-', '').strip()

        hits = by_en.get(nk, [])
        if not hits and ck:
            hits = by_cas.get(ck, [])
        target = None
        if hits:
            for i in hits:
                if COMPOUNDS[i].get('src') != '发酵食品文献库':
                    target = COMPOUNDS[i]; break
            if target is None:
                target = COMPOUNDS[hits[0]]

        if target:
            # 覆盖语义：优先用文件数据
            # 中文名：仅当文件给出"完整中文名"(≥3中文字)才覆盖，避免残缺碎片(如"苯酚")覆盖已有好名
            if cn and len(re.findall(r'[一-鿿]', cn)) >= 3:
                target['cn'] = cn
            if cat and cat != '其他':
                target['cat'] = cat                      # 覆盖类别
            if thr:
                target['thr'] = thr                     # 覆盖阈值
                target['med'] = med if med else target.get('med', '—')
            elif med:
                target['med'] = med
            if note:
                target['note'] = note                   # 覆盖风味描述（文件优先）
            target['src'] = '发酵食品文献库'
            for s in (x.get('syn') or []):
                if s and s not in (target.get('syn') or []):
                    target.setdefault('syn', []).append(s)
            updated += 1
        else:
            COMPOUNDS.append({
                'en': en, 'cn': cn if cn else en,        # 待核无中文：用英文名可读回退
                'cat': cat, 'thr': thr if thr else '—',
                'med': med if med else '文献未列',
                'note': note if note else '—',
                'src': '发酵食品文献库', 'syn': x.get('syn') or [],
                'cas': cas, '_ferment_pending': x.get('pending', False),
            })
            added += 1

    if added or updated:
        print('[info] fermented_food: 新增 %d, 覆盖 %d, 合计处理 %d' % (added, updated, len(items)))


def _load_fermented_food2():
    """合并 fermented_food_db2.json（发酵食品挥发性物质文献数据库 (1).xlsx，修订扩充版）。
    用户要求优先用该文件内容覆盖已有条目。该文件较旧版更完整（686→1249 键），含大量
    新增物质，但其中混入 191 条"待核化合物系统命名见英文"占位中文名、60 条以"-"开头的
    截断脏英文名。覆盖策略：
      - 中文名：仅当文件给出"真中文名"(≥3中文字且非待核/系统命名占位)才覆盖，防止把
        已有的好中文名(如"愈创木酚")覆盖成垃圾。
      - 脏英文名(以"-"开头)与占位中文名：作为新物质追加，并标 _ferment_pending 放行删除。
      - 类别/阈值/介质/风味描述：文件优先覆盖。
    """
    p = os.path.join(HERE, 'fermented_food_db2.json')
    if not os.path.exists(p):
        return
    try:
        items = json.load(open(p, 'r', encoding='utf-8'))
    except Exception as e:
        print('[warn] 读取 fermented_food_db2.json 失败:', e)
        return

    by_en, by_cas = {}, {}
    for i, c in enumerate(COMPOUNDS):
        k = normalize(c.get('en', ''))
        if k:
            by_en.setdefault(k, []).append(i)
        ck = (c.get('cas') or '').lower().replace('-', '').strip()
        if ck:
            by_cas.setdefault(ck, []).append(i)

    def _is_real_cn(s):
        """真正可用的中文名：含≥3中文字且不是待核/系统命名占位。"""
        s = (s or '').strip()
        if not s:
            return False
        if '待核' in s or '系统命名' in s:
            return False
        return len(re.findall(r'[一-鿿]', s)) >= 3

    added, updated = 0, 0
    for x in items:
        en = (x.get('en') or '').strip()
        cn = (x.get('cn') or '').strip()
        cas = (x.get('cas') or '').strip()
        cat = (x.get('cat') or '其他').strip()
        thr = (x.get('thr') or '').strip()
        med = (x.get('med') or '').strip()
        note = (x.get('note') or '').strip()
        nk = normalize(en)
        ck = cas.lower().replace('-', '').strip()
        # 脏英文名(以"-"开头)无法匹配现有条目 -> 直接作为新增追加
        dirty_en = en.startswith('-') or bool(re.search(r'[一-鿿]', en))

        target = None
        if not dirty_en:
            hits = by_en.get(nk, [])
            if not hits and ck:
                hits = by_cas.get(ck, [])
            if hits:
                for i in hits:
                    if COMPOUNDS[i].get('src') != '发酵食品文献库(1)':
                        target = COMPOUNDS[i]; break
                if target is None:
                    target = COMPOUNDS[hits[0]]

        if target:
            if _is_real_cn(cn):
                target['cn'] = cn
            if cat and cat != '其他':
                target['cat'] = cat
            if thr:
                target['thr'] = thr
                target['med'] = med if med else target.get('med', '—')
            elif med:
                target['med'] = med
            if note:
                target['note'] = note
            target['src'] = '发酵食品文献库(1)'
            prod = (x.get('products') or '').strip()
            if prod:
                target['products'] = prod
            for s in (x.get('syn') or []):
                if s and s not in (target.get('syn') or []):
                    target.setdefault('syn', []).append(s)
            updated += 1
        else:
            COMPOUNDS.append({
                'en': en, 'cn': cn if _is_real_cn(cn) else en,
                'cat': cat, 'thr': thr if thr else '—',
                'med': med if med else '文献未列',
                'note': note if note else '—',
                'src': '发酵食品文献库(1)', 'syn': x.get('syn') or [],
                'cas': cas, 'products': (x.get('products') or '').strip(),
                '_ferment_pending': (dirty_en or not _is_real_cn(cn)),
            })
            added += 1

    if added or updated:
        print('[info] fermented_food2: 新增 %d, 覆盖 %d, 合计处理 %d' % (added, updated, len(items)))


# ----------------------------------------------------------------------------
# 阈值单位统一折算为 μg/L（便于 OAV / ROAV 计算）
#   基质按 1 kg ≈ 1 L 近似：μg/kg → 1，mg/kg → 1000，μg/g → 1000；
#   水相体积浓度：1 ppb ≈ 1 μg/L，1 ppm = 1 mg/L = 1000 μg/L，1 ppt = 0.001 μg/L；
#   1 mg/m³ = 1 μg/L（严格等值）。
#   空气中的 ppb/ppm/ppt 为体积比(v/v)，缺分子量无法折算为 μg/L，沿用库内既有约定
#   「量纲不同未计入 OAV」，不参与计算。
# ----------------------------------------------------------------------------
THR_UNIT_TO_UGL = {
    'μg/l': 1.0, 'ug/l': 1.0, 'µg/l': 1.0,
    'ng/l': 0.001, 'mg/l': 1000.0, 'g/l': 1e6,
    'μg/kg': 1.0, 'ug/kg': 1.0, 'µg/kg': 1.0,
    'ng/kg': 0.001, 'mg/kg': 1000.0, 'g/kg': 1e6,
    'μg/g': 1000.0, 'ug/g': 1000.0, 'mg/g': 1e6,
    'mg/m³': 1.0, 'mg/m3': 1.0,
    'ppb': 1.0, 'ppm': 1000.0, 'ppt': 0.001,
}
_THR_UNIT_RE = re.compile(r'(μg|ug|µg|mg|ng|g)\s*/\s*(L|l|kg|g|m³|m3)|ppb|ppt|ppm', re.I)
_THR_NUM = r'\d[\d,]*(?:\.\d+)?'
_THR_RANGE_RE = re.compile(r'(' + _THR_NUM + r')\s*(?:[–—~～-]\s*(' + _THR_NUM + r'))?\s*(\S{0,4})$')
_THR_MED_KW = ('水', '空气', '食品基质', '基质', '乙醇', '体系', '蚝油', '汤', '肉', '奶', '油')


def _thr_med_of(seg):
    """取单个阈值分段的介质：优先全角括号内的介质描述，其次分段正文里的介质关键词。"""
    for m in re.finditer(r'（([^）]*)）', seg):
        s = m.group(1).strip()
        if any(k in s for k in _THR_MED_KW) and len(s) <= 30:
            return s
    core = re.sub(r'（[^）]*）', '', seg)
    for k in _THR_MED_KW:                       # 如「肉中 55–73 μg/kg」-> 肉中
        if k in core:
            return k + '中' if k in ('肉', '汤', '奶', '油') else k
    return '—'


# 上标数字 → 普通数字（处理 1.0×10⁵ 这类科学计数法）
_SUP = {'⁰': '0', '¹': '1', '²': '2', '³': '3', '⁴': '4', '⁵': '5',
        '⁶': '6', '⁷': '7', '⁸': '8', '⁹': '9'}


def _desup(s):
    return ''.join(_SUP.get(ch, ch) for ch in s)


def convert_threshold_ugl(raw):
    """把阈值文本折算为 μg/L，返回 (显示串, 数值或 None, 介质, 状态)。

    取值规则：区间取下限、多段(；分隔)取最小值 —— 属保守估计，使 OAV 偏大，
    用于关键致香物筛查时不会漏判。
    空气中的 v/v 单位【先剔除再取最小值】，否则像
    「0.007 ppb（空气）；0.021 μg/L（水）」会因选中空气值而把整条阈值丢弃。
    无显式单位的纯数字（如 Van Gemert / FTDB 的嗅觉阈值）按其固有 μg/L 处理。"""
    t = (raw or '').strip()
    if not t or t in ('—', '-', '–'):
        return ('—', None, '—', 'placeholder')
    # 预处理：上标科学计数法 ×10ⁿ → eⁿ
    t = _desup(t)
    t = re.sub(r'[×*xX]\s*10\s*([0-9]+)', lambda m: 'e' + m.group(1), t)
    cands = []
    for seg in re.split(r'[；;]', t):                 # 多段：mg/L；μg/L；mg/kg
        med = _thr_med_of(seg)
        core = re.sub(r'（[^）]*）', '', seg)          # 去掉括号注释后再解析
        for um in _THR_UNIT_RE.finditer(core):
            u = um.group(0).lower().replace(' ', '')
            if u not in THR_UNIT_TO_UGL:
                continue
            m = _THR_RANGE_RE.search(core[:um.start()].rstrip())
            if not m:
                continue
            lo = float(m.group(1).replace(',', ''))
            cands.append((lo * THR_UNIT_TO_UGL[u], u, med))
    if not cands:
        # 无显式单位：按 μg/L 解析纯数字（支持区间取下限、科学计数）
        for seg in re.split(r'[；;]', t):
            med = _thr_med_of(seg)
            core = re.sub(r'（[^）]*）', '', seg)
            m = re.search(r'(\d+(?:\.\d+)?(?:e[+-]?\d+)?)', core)
            if m:
                try:
                    lo = float(m.group(1))
                    cands.append((lo, 'μg/l', med))
                except ValueError:
                    pass
    if not cands:
        return ('—', None, _thr_med_of(t), 'unparseable')
    usable = [c for c in cands if not ('空气' in c[2] and c[1] in ('ppb', 'ppm', 'ppt'))]
    if not usable:
        return ('—', None, cands[0][2] + '（v/v，量纲不同未计入OAV）', 'air_vv')
    lo, u, med = min(usable, key=lambda x: x[0])
    return ('%g' % lo, lo, med, 'ok')


def _load_flavor_database_csv():
    """全量并入 flavor_database.csv —— 不做任何去重 / 合并 / 覆盖。

    列: category, subcategory, chemical_type, name_cn, name_en, cas, threshold, flavor, source

    用户要求：
      1) 文件中每一行都作为独立条目加入库，即使与库中已有物质相同也照加（不去重、不覆盖）；
      2) 物质分类【完全按文件】：chemical_type（发酵食品章节 342 行）优先，为空时取
         subcategory（其余章节 104 行），并去掉「1.2 」这类章节编号；
      3) 阈值统一折算为 μg/L（thr 显示 / thr_num 数值），便于后续 OAV、ROAV 计算。
    CSV 条目插到 COMPOUNDS 前部，使其在检索索引中优先命中（沿用"该文件数据优先"）。"""
    p = os.path.join(HERE, 'flavor_database.csv')
    if not os.path.exists(p):
        return
    try:
        import csv as _csv
        with open(p, 'r', encoding='utf-8-sig') as _f:
            rows = list(_csv.DictReader(_f))
    except Exception as e:
        print('[warn] 读取 flavor_database.csv 失败:', e)
        return

    def _cls(r):
        """文件给出的物质分类：chemical_type 优先，其次 subcategory（去掉章节编号）。"""
        for col in ('chemical_type', 'subcategory'):
            v = (r.get(col) or '').strip()
            if v and v != '—':
                return re.sub(r'^\d+(?:\.\d+)*\s*', '', v).strip() or '其他'
        return '其他'

    entries = []
    for r in rows:
        raw_en = (r.get('name_en') or '').strip()
        if not raw_en:
            continue
        # 区分两种括号：
        #  A) 前缀构型括号 "(E)-2-Nonenal" / "(E,Z)-2,4-Decadienal" -> 主体名 2-Nonenal，
        #     把带构型的全名补进同义名（匹配报告里的 (E)-... 写法）；
        #  B) 末尾别名括号 "2-Furfurylthiol (FFT)" -> 别名 FFT。
        #  不要把整行构型前缀当别名括号清空（否则 20 个合法化合物会被漏加）。
        syn = []
        m_pre = re.match(r'^\(([^)]*)\)\s*-?\s*(.+)$', raw_en)
        if m_pre:
            en = m_pre.group(2).strip()
            syn.append(raw_en)
        else:
            m_suf = re.search(r'\(([^)]*)\)', raw_en)
            if m_suf:
                en = raw_en[:m_suf.start()].strip()
                a = m_suf.group(1).strip()
                if a:
                    syn.append(a)
            else:
                en = raw_en
        if not en:
            continue
        cn = (r.get('name_cn') or '').strip()
        thr, thr_num, med, _st = convert_threshold_ugl(r.get('threshold'))
        # 该文件有 63 行 name_cn 列填的是英文名；保留原样并补进同义名，保证仍可检索
        if cn and cn != en and not re.search(r'[一-鿿]', cn) and cn not in syn:
            syn.append(cn)
        entries.append({
            'en': en,
            'cn': cn if cn else en,
            'cat': _cls(r),
            'thr': thr,
            'thr_num': thr_num,                    # μg/L 数值，供 OAV / ROAV 计算
            'thr_raw': (r.get('threshold') or '').strip(),
            'med': med,
            'note': (r.get('flavor') or '').strip() or '—',
            'src': '风味数据库CSV',
            'cas': (r.get('cas') or '').strip(),
            'syn': syn,
            'products': (r.get('source') or '').strip(),
            '_csv': True,                          # 标记：后续清洗不去重、不重分类
        })

    if entries:
        COMPOUNDS[:0] = entries                    # 置于前部 -> 检索索引优先命中
        n_num = sum(1 for e in entries if e.get('thr_num'))
        print('[info] flavor_database.csv: 全量加入 %d 条（不去重）；物质分类按文件；'
              '阈值已折算 μg/L（其中 %d 条可直接用于 OAV 计算）' % (len(entries), n_num))


_load_custom()
_load_ftdb()
_load_curated_extras()
_load_volatile_html()
_load_fermented_food()
_load_fermented_food2()
_load_flavor_database_csv()
_drop_column_bleed()
_drop_artifacts()
_clean_notes()
_expand_categories()
_normalize_categories()
_fix_display()
_clean_cn_references()
_clean_cn_spaces()
_fix_pyrazine_cn()
_cn_from_cas()
_fix_frag_cn()
_fix_ester_lactone_cn()
_dedupe_ester_lactone()
_normalize_thresholds_ugl()
_drop_no_cn()
_drop_unknown()
_dedupe_odor()
_finalize_columns()
_dedupe_pyrazine_cas()
_clean_pyrazine_syns()
_dedupe_cas_global()
DB = _build_index()


# ----------------------------------------------------------------------------
# 5. 匹配引擎
# ----------------------------------------------------------------------------
def _build_norm_boundary(name):
    """返回 (归一化串 nq, b4)：b4[i]=True 表示 nq[i] 之前是分隔符/行首（即词边界）。
    分隔符（空格/括号/连字符/逗号等）被剥离，但会标记其后字符为词边界，用于整词匹配。"""
    nq = []
    b4 = []
    prev_sep = True
    for ch in (name or ''):
        lc = ch.lower()
        if re.match(r'[a-z0-9一-鿿]', lc):
            nq.append(lc)
            b4.append(prev_sep)
            prev_sep = False
        else:
            prev_sep = True
    return ''.join(nq), b4


def _boundary_hit(nq, b4, k):
    """k 是否在 nq 中以「整词」出现（前后均为词边界）。
    返回 'start'（词首边界，优先）或 'end'（词尾边界），否则 None。"""
    if not k:
        return None
    L = len(k)
    N = len(nq)
    i = nq.find(k)
    while i >= 0:
        if b4[i] and (i + L >= N or b4[i + L]):
            return 'start' if b4[i] else 'end'
        i = nq.find(k, i + 1)
    return None


def _match_candidate(q, nq, b4, name, fuzzy):
    """对单个归一化候选查询做 精确 / 整词边界包含 / 模糊 匹配，命中返回 dict，否则 None。

    整词边界包含：库化合物名须以词边界（空格/括号/连字符/首尾）出现在原物质名中，
    且优先词首匹配——从而杜绝 'ethyl' 藏在 'methyl…'、'benzaldehyde' 附在 'vanillin…' 后的贪婪子串误匹配。"""
    if not q:
        return None
    if q in DB:
        return dict(DB[q], match='精确', score=1.0, query=name)
    pre_k = None   # 词首整词匹配（最长键优先）
    suf_k = None   # 词尾整词匹配（最长键优先）
    for k, c in DB.items():
        if not k or len(k) < 3:
            continue
        hit = _boundary_hit(nq, b4, k)
        if hit == 'start':
            if pre_k is None or len(k) > len(pre_k):
                pre_k = k
        elif hit == 'end':
            if suf_k is None or len(k) > len(suf_k):
                suf_k = k
    if pre_k is not None:
        return dict(DB[pre_k], match='包含匹配', score=0.9, query=name)
    if suf_k is not None:
        return dict(DB[suf_k], match='包含匹配', score=0.9, query=name)
    if fuzzy:
        keys = list(DB.keys())
        close = difflib.get_close_matches(q, keys, n=1, cutoff=0.82)
        if close:
            c = DB[close[0]]
            sc = difflib.SequenceMatcher(None, q, close[0]).ratio()
            return dict(c, match='模糊匹配', score=round(sc, 2), query=name)
    return None


def match_compound(name, fuzzy=True):
    """多级匹配：精确 -> 边界包含 -> 模糊。未命中则按命名结构自动分类。

    解析名常带括号（如 'Butanoic acid (branched)'、'Pear ester (ethyl 2-methylbutyrate)'、
    'Orange oil (Limonene)'）：主名未命中或仅含杂质时，额外把「括号里的文字」与「去括号后的主体」
    作为候选去匹配数据库，以便拿到中文名与阈值，正确计算 OAV / ROVA。"""
    q = normalize(name)
    if not q:
        return dict(en=name, cn='(空)', cat='其他', thr='—', med='—',
                    note='输入为空', src='—', match='无效', score=0.0, query=name)
    # 基于原名的词边界信息（供整词包含匹配，区分 'ethyl' 藏在 'methyl…' 等）
    nq, b4 = _build_norm_boundary(name)
    # 候选查询：全名 -> 各括号片段 -> 去括号后的主体（去重）
    cands = [q]
    for seg in re.findall(r'\(([^()]*)\)', name or ''):
        qs = normalize(seg)
        if len(qs) >= 3:
            cands.append(qs)
    base_q = normalize(re.sub(r'\([^()]*\)', ' ', name or ''))
    if base_q and base_q != q:
        cands.append(base_q)
    seen = set()
    for cand in cands:
        if cand in seen:
            continue
        seen.add(cand)
        # 模糊匹配仅在全名候选上运行（控制开销）；括号/主体片段靠精确+整词包含即可覆盖绝大多数情形
        hit = _match_candidate(cand, nq, b4, name, fuzzy and cand == q)
        if hit:
            return hit
    # 未命中 -> 自动分类
    cat = auto_classify(name)
    return dict(en=name, cn='(未收录)', cat=cat, thr='—', med='—',
                note='未在数据库，已按命名结构自动推断类别；阈值请补充（可联网检索或写入 custom_compounds.json）',
                src='—', match='未匹配(自动分类)', score=0.0, query=name)


def enrich(rows):
    """rows: [{'en':..., 'rt':..., 'resp':..., 'rr':..., 'conc0':..., 'conc':..., 'istd_area':..., ...}]
    返回带中文/分类/阈值的 enriched 列表。"""
    out = []
    for r in rows:
        m = match_compound(r.get('en', ''))
        item = dict(m)
        item.update({k: r.get(k) for k in ('rt', 'resp', 'rr', 'conc', 'conc0', 'istd_area') if k in r})
        # 报告自带浓度可能只存在于 'conc' 字段，统一透传为 'conc0' 供前端回退显示
        if 'conc0' not in item and 'conc' in r:
            item['conc0'] = r.get('conc')
        out.append(item)
    return out


# ----------------------------------------------------------------------------
# 6. 解析 Agilent MassHunter 导出的定量报告
# ----------------------------------------------------------------------------
import openpyxl

def _cell_num(x):
    """把单元格值转成 float；'ND'/'—'/空/非数字返回 None。"""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).replace(',', '').replace('\n', ' ').replace('\r', ' ').strip()
    if s in ('', 'ND', 'nd', 'N.D.', 'N/A', 'na', '—', '-', '–', 'NULL', 'null', '未检出',
             'N.D', 'n.d.', 'LOD', 'lod', '检出限', 'trace', 'Trace'):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _find_header(rows):
    """在表格中定位表头行：含『化合物名』关键字且含至少一列数值关键字。
    返回 (行索引, 表头元组)。找不到返回 (None, None)。"""
    comp_kw = ('化合物', '化合物名称', '英文名', '名称', 'compound', 'name', 'comp.')
    val_kw = ('rt', '保留时间', '出峰', '响应', '峰面积', '响应比', '含量', 'area',
              'ratio', '浓度', 'final', 'istd', '内标')
    best = None
    for i, r in enumerate(rows):
        cells = [str(x).strip() for x in r if x is not None]
        if not cells:
            continue
        low = [c.lower() for c in cells]
        has_comp = any(any(k in c for k in comp_kw) for c in low)
        has_val = any(any(k in c for k in val_kw) for c in low)
        if has_comp and has_val:
            return i, r
        # 退化情形：仅含『化合物名』列也接受（后续行仍有数值）
        if has_comp and best is None:
            best = (i, r)
    return best if best else (None, None)


def _map_cols(header):
    """按关键字模糊映射各列下标。返回字典（缺列为 None）。"""
    hdr = [(str(x).strip() if x is not None else '') for x in header]
    low = [h.lower() for h in hdr]

    def first(*keys, exclude=(), include=()):
        for k in keys:
            for j, h in enumerate(low):
                if k in h and not any(e in h for e in exclude) \
                        and (not include or any(ic in h for ic in include)):
                    return j
        return None

    comp = first('化合物', '英文名', '名称', 'compound', 'name', 'comp.')
    rt = first('rt', '保留时间', '出峰')
    # 峰面积/响应值列：命中『峰面积/响应/area』但不含『内标/istd』
    area = first('峰面积', '响应值', '响应', 'area', exclude=('内标', 'istd'))
    # 响应比列：优先『响应比/相对响应/相对峰面积/含量(=响应比)/ratio』
    rr = first('响应比', '相对响应', '相对峰面积', 'ratio', exclude=('内标', 'istd'))
    if rr is None:
        rr = first('含量', exclude=('内标', 'istd', '最终浓度', '浓度'))
    # 浓度列
    conc = first('最终浓度', '浓度', exclude=('内标', 'istd', '响应比', '含量'))
    # 内标名 / 内标峰面积 两列
    istd_name = first('istd', '内标', exclude=('响应', '面积'))
    istd_area = first('istd', '内标', include=('响应', '面积'))
    # 兜底：若未单独分出内标峰面积，但存在『响应比』缺失而『内标』列是数字列，则忽略
    return {
        'comp': comp, 'rt': rt, 'area': area, 'rr': rr, 'conc': conc,
        'istd_name': istd_name, 'istd_area': istd_area,
    }


def parse_gcms_excel(path):
    """通用解析 GC-MS 定量报告（Agilent MassHunter 导出 / 已处理结果表 / 任意含物质列的表格）。

    返回 dict:
      {
        'rows':   [{'en','rt','resp','rr','conc','istd_area', ...}],
        'ist_name': 自动识别的内标物质名(str) 或 None,
        'ist_area': 自动识别的内标峰面积(float) 或 None,
        'schema': 'agilent' | 'result' | 'generic',
        'headers': [表头文本...],
      }
    特点:
      - 按关键字模糊定位表头行与各数值列，兼容多种表头（化合物/英文名/中文名、
        RT(min)/保留时间、峰面积响应(原始)/响应、响应比/含量(=响应比)、最终浓度/浓度）。
      - 响应比缺失时，用 化合物峰面积 / 内标峰面积 反算 rr。
      - 同一色谱峰（相同 RT+峰面积+响应比）的 MassHunter 多库匹配行自动去重，仅保留首条。
      - 自动识别内标物质名（ISTD 列）与内标峰面积（ISTD 响应 列）。
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    hidx, header = _find_header(rows)
    if hidx is None or header is None:
        return {'rows': [], 'ist_name': None, 'ist_area': None,
                'schema': 'unknown', 'headers': []}

    cols = _map_cols(header)
    if cols['comp'] is None:
        return {'rows': [], 'ist_name': None, 'ist_area': None,
                'schema': 'unknown', 'headers': [str(x).strip() for x in header if x is not None]}

    schema = 'agilent' if any('istd' in str(x).lower() for x in header if x) else \
        ('result' if any('含量' in str(x) or '响应比' in str(x) for x in header if x) else 'generic')

    SKIP = ('样品类型', '化合物', '定量分析完成报告', '样品色谱图', '序号', 'RT',
            '保留时间', 'name', 'compound', 'comp.', '英文名', '中文名', '类别', 'name ')

    def g(r, key):
        j = cols.get(key)
        if j is None or j >= len(r):
            return None
        return r[j]

    out = []
    seen_sig = set()
    ist_name = None
    ist_area = None
    for r in rows[hidx + 1:]:
        c0 = g(r, 'comp')
        comp = (str(c0).replace('\n', ' ').replace('\r', ' ').strip()
                if isinstance(c0, str) else (str(c0).strip() if c0 is not None else ''))
        if not comp or comp in SKIP:
            continue
        rt = _cell_num(g(r, 'rt'))
        resp = _cell_num(g(r, 'area'))
        rr = _cell_num(g(r, 'rr'))
        istdr = _cell_num(g(r, 'istd_area'))
        conc_cell = g(r, 'conc')
        # 内标名/峰面积识别（ISTD 列是字符串名；ISTD 响应 列是数字）
        if cols['istd_name'] is not None and cols['istd_name'] < len(r):
            v = r[cols['istd_name']]
            if isinstance(v, str) and v.strip() and ist_name is None:
                ist_name = v.strip()
        if istdr is not None and ist_area is None:
            ist_area = istdr
        # 响应比反算（缺 rr 时用 峰面积/内标峰面积）
        if rr is None and resp is not None and istdr:
            rr = resp / istdr
        if rt is None and resp is None and rr is None and conc_cell is None:
            continue
        # 同峰多库命中去重：相同 RT+峰面积+响应比 只保留首条
        sig = (round(rt or 0, 3), round(resp or 0, 0), round(rr or 0, 4))
        if sig in seen_sig:
            continue
        seen_sig.add(sig)
        item = {
            'en': comp,
            'rt': rt,
            'resp': int(resp) if resp is not None else None,
            'rr': rr,
            'conc': conc_cell,
            'istd_area': istdr,
        }
        out.append(item)
    return {
        'rows': out,
        'ist_name': ist_name,
        'ist_area': ist_area,
        'schema': schema,
        'headers': [str(x).strip() for x in header if x is not None],
    }


def parse_report_text(text, value_col=None, ist_name='2-辛醇'):
    """从 PDF / Markdown / 纯文本 GC-MS 报告中扫描并识别已知风味物质。

    按【表头语义】识别各数值列（去除固定栏位顺序的假设），将提取结果对应到
    内标法浓度公式的字母标识：
        A  = 化合物峰面积（响应 / 峰面积 / area）
        A₁ = 内标峰面积（ISTD / 内标 列，或从内标物质所在行提取）
        rr = 响应比（A / A₁，无量纲）—— 公式中直接使用 rr 计算浓度
        c  = 浓度（最终浓度 / 浓度，报告自带）
        RT = 保留时间
    计算时直接代入公式 C = (c₁ × V₁ × rr) / m，因 rr 已含 A₁，无需再除以内标峰面积。
    其中 c₁(内标浓度) / V₁(内标加量) / m(样品量) 由用户在界面填入。

    内标物质默认 '2-辛醇'（ist_name 可覆盖）；报告无某列时该字段为 None。
    """
    # 建立 归一化名 -> 标准英文名 的索引（优先英文，再中文/同义名）
    idx = {}
    for c in COMPOUNDS:
        names = [c.get('en', '')] + (c.get('syn') or [])
        if c.get('cn'):
            names.append(c.get('cn'))
        for nm in names:
            if not nm:
                continue
            nk = normalize(nm)
            # 剔除碎片候选，避免误命中：
            #   - 纯数字名直接跳过；
            #   - ASCII 名（英文/编号）要求 ≥3 字符，滤掉 '1'、'2'、'ol' 之类；
            #   - 中文名要求 ≥2 字，滤掉单字 '酸' '醛' '醇'，但保留「丁酸/己醛」等合法两字名。
            if not nk or nk.isdigit():
                continue
            if nk.isascii():
                if len(nk) < 3:
                    continue
            elif len(nk) < 2:
                continue
            if nk and nk not in idx:
                idx[nk] = c.get('en', '')
    if not idx:
        return {'rows': [], 'ist_name': ist_name, 'ist_area': None,
                'schema': 'text', 'headers': []}

    # 候选词按长度降序（先匹配长名，避免短名抢先）
    cands = sorted(idx.keys(), key=len, reverse=True)
    # 归一化并记录「归一化下标 -> 原文下标」的映射，便于回原文取数
    norm_text, pos_map = _normalize_with_map(text)
    if not norm_text:
        return []

    # 表义识别各列（返回 role -> 表头单元格 0-based 索引；ncols=表头单元格数）。
    # 与数据行按尾部对齐取数，避免「ISTD 响应」拆词、内标名列含数字、化合物名含空格导致错位。
    cols, ncols = _detect_columns(text)
    rt_i = cols.get('rt'); area_i = cols.get('area')
    rr_i = cols.get('rr'); conc_i = cols.get('conc'); istd_i = cols.get('istd')
    # 按表头单元格索引从数据行取第 idx 列的数值。
    # 数据行比表头多出的 token（化合物名含空格）统一算到行首偏移，不影响尾部数值列对齐。
    def _val_at(line, idx):
        if idx is None or ncols <= 0:
            return None
        toks = line.split()
        off = len(toks) - ncols
        if off < 0:
            off = 0
        ti = idx + off
        if 0 <= ti < len(toks):
            s = toks[ti].strip()
            if re.match(r'^-?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?$', s):
                return float(s)
        return None
    # A₁ 内标峰面积：优先从专用 istd 列取（主循环里首个匹配物质行获得，通常全表恒定）
    istd_area = None

    found = []          # [{'en':..., 'rr':响应比, 'conc0':原始浓度}]
    seen_en = set()
    taken = []          # 已被更长候选占用的区间 (start, end)，避免短名嵌套命中（如「乙酸」命中「乙酸乙酯」）

    def _overlaps(a, b):
        return not (b[1] <= a[0] or b[0] >= a[1])


    def _at_word_boundary(text, o_s, o_e):
        """判断原文 [o_s, o_e] 是否处于词边界：仅当前后紧邻字母/数字/中文时拒
        （避免「芳樟醇」误嵌在「反式-呋喃型氧化芳樟醇」中而错配上一行的值）。
        注意：连字符/括号/逗号等已被归一化剥离，属同名的内部分隔符，不能算作非边界
        ——否则会误杀「β-月桂烯」「D-柠檬烯」等行首英文名。行首/空格/换行仍天然为边界。"""
        if o_s > 0:
            prev = text[o_s - 1]
            if prev.isalnum() or '\u4e00' <= prev <= '\u9fff':
                return False
        if o_e + 1 < len(text):
            nxt = text[o_e + 1]
            if nxt.isalnum() or '\u4e00' <= nxt <= '\u9fff':
                return False
        return True

    for nk in cands:
        en = idx[nk]
        if en in seen_en:
            continue
        # 找第一个「不与已占用区间重叠」且「处于词边界」的出现位置
        start, pos = 0, -1
        while True:
            p = norm_text.find(nk, start)
            if p < 0:
                break
            o_s = pos_map[p]; o_e = pos_map[p + len(nk) - 1]
            if _at_word_boundary(text, o_s, o_e) and \
               not any(_overlaps((p, p + len(nk)), iv) for iv in taken):
                pos = p
                break
            start = p + 1
        if pos < 0:
            continue
        taken.append((pos, pos + len(nk)))
        seen_en.add(en)
        s = pos_map[pos + len(nk) - 1] + 1
        # 取该物质所在的完整数据行（含全部数值列）
        line_start = text.rfind('\n', 0, s) + 1
        e = text.find('\n', s)
        line = text[line_start: e if e >= 0 else len(text)]
        rt = _val_at(line, rt_i)
        resp = _val_at(line, area_i)     # A 化合物峰面积（响应）
        rr = _val_at(line, rr_i)         # rr 响应比
        conc0 = _val_at(line, conc_i)    # c 浓度（报告自带）
        found.append({'en': en, 'rt': rt, 'resp': resp, 'rr': rr, 'conc0': conc0})
        if istd_area is None and istd_i is not None:
            istd_area = _val_at(line, istd_i)
    # A₁ 内标峰面积挂到每行（Python list 不可挂属性，存入 dict 以便 JSON 序列化）
    if istd_area is not None:
        for r in found:
            r['istd_area'] = istd_area
    return {'rows': found, 'ist_name': ist_name, 'ist_area': istd_area,
            'schema': 'text', 'headers': []}


# ----------------------------------------------------------------------------
# 8b. 解析 Shimadzu GCMSsolution 多处理组定量报告（"定量分析完成报告"）
#     自动识别数据文件 / 样品名称（处理组），提取每个化合物在各处理组的
#     RT / 响应 / ISTD响应 / 响应比 / 最终浓度。一次解析即可得到「处理组对比」所需全部数据。
# ----------------------------------------------------------------------------
def parse_gcms_pdf_multi(path, ist_name=None):
    """解析岛津 GCMSsolution 多处理组定量报告（PDF）。

    通过 pdftotext -layout 保留「数据文件 | 化合物 | ISTD | 样品类型 | RT | 响应 |
    ISTD响应 | 响应比 | 最终浓度」的表格结构，逐块提取：
      - 处理组：由顶部「数据文件 / 样品名称」表自动识别（如 1.d→1, 3.d→3, ck.d→ck）；
      - 每个化合物在每个处理组的 RT / 响应 / ISTD响应 / 响应比 / 最终浓度。
    内标法浓度可直接用 响应比 rr（= 化合物峰面积 ÷ ISTD 峰面积）按
    C = c₁·V₁·rr ÷ m 计算，故每个处理组各自携带 rr，无需再除以内标峰面积。

    返回结构:
      {'multi': True, 'schema': 'shimadzu_quant_multi', 'groups':[组名...],
       'samples':[数据文件名...], 'ist_name': 内标名, 'ist_area_by_group':{组:内标峰面积},
       'n_compounds': N, 'compounds':[{en,cn,cat,thr,...,'by_group':{组:{...}}}],
       'rows':[展平后的逐化合物逐处理组行]}
    未识别到定量结果时返回 {'multi': True, 'compounds': [], 'groups': [], 'samples': []}。
    """
    import subprocess
    try:
        txt = subprocess.run(['pdftotext', '-layout', '-f', '1', '-l', '324', path, '-'],
                              capture_output=True, text=True, timeout=120).stdout
    except Exception:
        txt = ''
    if not txt:
        return {'multi': True, 'compounds': [], 'groups': [], 'samples': [],
                'schema': 'shimadzu_quant_multi'}
    lines = txt.split('\n')

    def _num(x):
        x = (x or '').replace(',', '').strip()
        if x in ('ND', 'nd', 'N.D.', 'N/A', '—', '-', '–', ''):
            return None
        try:
            return float(x)
        except (ValueError, TypeError):
            return None

    SAMPLE_RE = re.compile(r'^([A-Za-z0-9_.\-]+\.d)\b')
    sample_to_name = {}
    samples = []
    ist_name_detected = None
    ist_area_by_group = {}
    # piv: 化合物标准名 en -> {'en','cn','cat','thr','cas','source','match','odor','by_group':{组:{...}}}
    piv = {}
    cur = None
    frags = []
    header_open = False
    in_quant = False

    for line in lines:
        s = line.strip()
        if not s:
            continue
        if '定量结果' in s and '数据文件' not in s:
            in_quant = True
            continue
        # 顶部「数据文件 / 样品名称」表：自动识别处理组
        m = SAMPLE_RE.match(line)
        if not in_quant and m:
            samp = m.group(1)
            toks = line.split()
            if len(toks) >= 2 and samp == toks[0] and samp not in sample_to_name:
                sample_to_name[samp] = toks[1]
                samples.append(samp)
            continue
        if not in_quant:
            continue
        # 子表头行（数据文件 | 化合物 | ISTD ...）：封口当前化合物名
        if s.startswith('数据文件') and '化合物' in s and 'ISTD' in s:
            if header_open and frags:
                cur = ' '.join(frags).strip()
                frags = []
            header_open = False
            continue
        m = SAMPLE_RE.match(line)
        if m:
            samp = m.group(1)
            if samp not in samples:
                samples.append(samp)
            if header_open and frags:
                cur = ' '.join(frags).strip()
                frags = []
                header_open = False
            toks = line.split()
            # 内标名 = Sample 前一词（ISTD 列）
            if 'Sample' in toks:
                i = toks.index('Sample')
                if i >= 1 and ist_name_detected is None:
                    cand = toks[i - 1]
                    if cand and cand not in ('化合物', 'ISTD'):
                        ist_name_detected = cand
                nums = [toks[i + 1:i + 6]]
            else:
                nums = [None]
            nums = nums[0] if isinstance(nums[0], list) else nums
            nums = [_num(t) for t in nums] if isinstance(nums, list) else [None]
            rec = {}
            for k, v in zip(['rt', 'resp', 'istd_resp', 'rr', 'final_conc'], nums):
                rec[k] = v
            rec['sample'] = samp
            # 归并到标准名
            mcomp = match_compound(cur or samp)
            en = mcomp.get('en') or cur or samp
            entry = piv.get(en)
            if entry is None:
                entry = {
                    'en': en, 'cn': mcomp.get('cn'), 'cat': mcomp.get('cat', '其他'),
                    'thr': mcomp.get('thr'), 'cas': mcomp.get('cas'),
                    'source': mcomp.get('source'), 'match': mcomp.get('match'),
                    'odor': mcomp.get('odor'), 'by_group': {},
                }
                piv[en] = entry
            grp = sample_to_name.get(samp, samp)
            if grp not in entry['by_group'] or entry['by_group'][grp] is None:
                entry['by_group'][grp] = {k: rec.get(k) for k in
                                          ('rt', 'resp', 'istd_resp', 'rr', 'final_conc', 'sample')}
            if rec.get('istd_resp') is not None and grp not in ist_area_by_group:
                ist_area_by_group[grp] = rec['istd_resp']
            continue
        # 缩进行：子表头前的化合物名折行（其余为化合物列折行，忽略）
        if line[:1] in ' \t':
            if header_open:
                frags.append(s)
            continue
        # 新化合物名（列 0）
        if header_open and frags:
            cur = ' '.join(frags).strip()
        frags = [s]
        header_open = True

    compounds = list(piv.values())
    groups = [sample_to_name.get(s, s) for s in samples]
    if not compounds:
        return {'multi': True, 'compounds': [], 'groups': groups, 'samples': samples,
                'schema': 'shimadzu_quant_multi'}
    # 展平为逐化合物逐处理组行
    rows = []
    sample_by_group = {sample_to_name.get(s, s): s for s in samples}
    for c in compounds:
        for g in groups:
            cell = c['by_group'].get(g)
            if not cell:
                continue
            row = {k: c.get(k) for k in
                   ('en', 'cn', 'cat', 'thr', 'cas', 'source', 'match', 'odor')}
            row['group'] = g
            row['sample'] = sample_by_group.get(g, g)
            row.update(cell)
            rows.append(row)
    return {
        'multi': True, 'schema': 'shimadzu_quant_multi',
        'software': 'Shimadzu GCMSsolution 定量分析完成报告',
        'groups': groups, 'samples': samples,
        'ist_name': ist_name or ist_name_detected,
        'ist_area_by_group': ist_area_by_group,
        'n_compounds': len(compounds),
        'compounds': compounds, 'rows': rows,
    }


# ---- 数值提取 ----
# 报告中可能出现的数值列标签（按优先级），用于避开 RT（保留时间）等干扰列
_RR_LABELS = ('响应比', '响应', '含量', '浓度', 'ratio', 'resp', 'conc')
# 表头里用于锁定「响应比/含量」列的标签
_RR_HEAD = ('响应比', '含量', '浓度', 'ratio', 'conc', '响应', 'resp')
# 表头里明显不是数值列的格子（单位、名称、定性等）
_NON_VALUE_HEAD = ('单位', 'unit', '名称', 'name', 'cas', '定性', 'qual', 'compound', '化合物', 'ion', '离子')


def _line_at(text, pos):
    """返回 pos 所在的那一行文本。"""
    s = text.rfind('\n', 0, pos) + 1
    e = text.find('\n', pos)
    return text[s: e if e >= 0 else len(text)]


def _line_tail(text, pos):
    """返回 pos 起到本行末尾的文本。
    不含物质名本身，避免把「2-庚酮」「3-甲基丁醛」等名字里的数字当成数值。"""
    e = text.find('\n', pos)
    return text[pos: e if e >= 0 else len(text)]


def _split_header(line):
    """把表头行拆成单元格；合并 'ISTD 响应' 这类被空白拆开的两词表头为单个单元格。"""
    cells = [c.strip() for c in re.split(r'\s+|\|', line.strip()) if c.strip()]
    merged = []
    i = 0
    while i < len(cells):
        c = cells[i]
        cl = c.lower()
        # 「ISTD/内标」后紧跟「响应/response/面积」=> 这是「内标峰面积」列，合并成一格
        if cl in ('istd', '内标') and i + 1 < len(cells):
            nxt = cells[i + 1].lower()
            if any(k in nxt for k in ('响应', 'response', '面积')):
                merged.append(c + ' ' + cells[i + 1])
                i += 2
                continue
        merged.append(c)
        i += 1
    return merged


def _detect_columns(text):
    """按表义识别表头各列。

    返回 (pos, ncols)：pos={role: 0-based 表头单元格索引(含名称列)}；ncols=表头单元格总数。
    列识别基于表头单元格（已合并 'ISTD 响应'），与数据行按尾部对齐取数，
    从而不受化合物名含空格、或 'ISTD 响应' 拆词的影响。

    关键修正：istd 只匹配「内标峰面积」类数值列，不再含裸 'istd'/'内标'——
    否则会误命中 ISTD 名称列（其值如 '2-辛醇' 含数字，会被当成 A₁ 填入响应值）。
    """
    # 公式字母 -> 表头候选词。istd 置于 area 之前，且不含裸 'istd'/'内标'。
    specs = {
        'rt':   ('保留时间', 'retention', 'rt', '出峰时间', 'time'),
        'istd': ('istd响应', 'istd 响应', '内标峰面积', '内标响应', 'istd面积', 'istd 面积'),
        'rr':   ('响应比', '相对响应', '相对峰面积', 'ratio', 'rr'),
        'conc': ('最终浓度', '浓度', '含量', 'conc', 'content'),
        'area': ('峰面积', '响应值', '响应', '面积', 'area', 'peak', 'response'),
    }
    for line in text.split('\n')[:80]:
        low = line.lower()
        if not ('化合物' in low or 'compound' in low or '物质' in low):
            continue
        if not any(lb in low for lb in ('rt', '响应', '面积', '浓度', '峰', 'ratio', 'retention')):
            continue
        cells = _split_header(line)
        if len(cells) < 3:
            continue
        name_i = next((j for j, c in enumerate(cells) if c.lower() in ('化合物', 'compound', '物质')), 0)
        pos = {}
        for j in range(len(cells)):
            if j == name_i:
                continue
            c = cells[j].lower()
            if any(u in c for u in _NON_VALUE_HEAD):
                continue
            for f, lbs in specs.items():
                if any(lb in c for lb in lbs):
                    if f not in pos:
                        pos[f] = j
                    break
        if pos:
            return pos, len(cells)
    return {}, 0


def _extract_num_at_col(text, start, col_pos):
    """从原文 start 处所在行起，取第 col_pos 个数值（1-based）。
    表格被换行拆开时向下多看一行；取不到返回 None。"""
    import re
    if not col_pos or col_pos < 1:
        return None
    seg = _line_tail(text, start)
    nums = re.findall(r'\d+\.?\d*', seg)
    if len(nums) < col_pos:
        e = text.find('\n', start)
        if e >= 0:
            nums = (nums + re.findall(r'\d+\.?\d*', _line_at(text, e + 1)))[:24]
    if len(nums) >= col_pos:
        try:
            return float(nums[col_pos - 1])
        except ValueError:
            return None
    return None


def _normalize_with_map(text):
    """归一化并返回 (归一化串, 位置映射表)。
    映射表 map[i] = 归一化串第 i 个字符在原文中的下标，用于把命中位置还原到原文。"""
    if not isinstance(text, str):
        return '', []
    chars, pos = [], []
    for i, ch in enumerate(text):
        c = ch.lower().replace('\u2019', "'").replace('\u2018', "'")
        if len(c) == 1 and (('a' <= c <= 'z') or ('0' <= c <= '9') or '\u4e00' <= c <= '\u9fff'):
            chars.append(c)
            pos.append(i)
    return ''.join(chars), pos

# ----------------------------------------------------------------------------
# 7. 便捷：分类汇总
# ----------------------------------------------------------------------------
# ---- 信息完整度：优先展示「有风味描述 + 有阈值 + 有来源」的物质 ----
# 阈值区分度最大（库内仅约 62% 有值），故权重更高；风味描述、来源各 1 分。
_FIELDS_WEIGHT = (('odor', 1), ('thr', 2), ('source', 1))

def has_field(c, key):
    """判断某字段是否真正有值（排除空串与占位符）。"""
    v = c.get(key)
    if v is None:
        return False
    s = str(v).strip()
    return s not in ('', '—', '-', '–', '暂无', '无', 'nan', 'None', 'null')

def completeness(c):
    """信息完整度得分：满分 4（风味描述1 + 阈值2 + 来源1）。"""
    return sum(w for k, w in _FIELDS_WEIGHT if has_field(c, k))

def completeness_label(c):
    """完整度文本标签，便于界面标注。"""
    s = completeness(c)
    if s >= 4:
        return '信息完整'
    if s >= 2:
        return '部分缺失'
    return '信息待补'

def sort_by_completeness(rows):
    """按信息完整度降序：有风味描述/阈值/来源的物质优先展示。
    同分时保持原有顺序（稳定排序），不破坏报告原有物质次序。"""
    return sorted(rows, key=lambda r: -completeness(r))

def category_counts(results):
    from collections import Counter
    return dict(Counter(r['cat'] for r in results))


def extract_substance_names(text):
    """从任意自然语言/表格文本中提取库内已知风味物质的标准英文名列表（去重、保序）。
    供批量检索的文件解析复用：扫描全文（归一化、去空格）按子串匹配库内物质。"""
    import re
    idx = {}
    for c in COMPOUNDS:
        names = [c.get('en', '')] + (c.get('syn') or [])
        if c.get('cn'):
            names.append(c.get('cn'))
        for nm in names:
            if not nm:
                continue
            nk = normalize(nm)
            if len(nk) < 3:
                continue
            if nk and nk not in idx:
                idx[nk] = c.get('en', '')
    if not idx:
        return []
    cands = sorted(idx.keys(), key=len, reverse=True)
    norm_text = normalize(text)
    out, seen = [], set()
    for nk in cands:
        en = idx[nk]
        if en in seen:
            continue
        if norm_text.find(nk) >= 0:
            seen.add(en)
            out.append(en)
    return out


# ----------------------------------------------------------------------------
# 8. 阈值解析 + 浓度单位换算 + OAV / ROVA 计算
#    阈值(odor threshold)统一单位 μg/kg(水相近似)。
#    浓度换算: 把内标法算出的浓度 C（单位取决于输入单位组合）折算到 μg/kg，
#              以便与阈值直接相除得到 OAV(气味活性值)。
# ----------------------------------------------------------------------------
# 浓度单位换算基础因子（质量 -> 克，体积 -> 毫升）
_MASS_TO_G = {
    'pg': 1e-12, 'ng': 1e-9, 'µg': 1e-6, 'μg': 1e-6, 'ug': 1e-6,
    'mg': 1e-3, 'g': 1.0, 'kg': 1e3,
    'NG': 1e-9, 'UG': 1e-6, 'UG': 1e-6, 'MG': 1e-3, 'G': 1.0, 'KG': 1e3,
}
_VOL_TO_ML = {
    'pl': 1e-12, 'nl': 1e-9, 'µl': 1e-3, 'μl': 1e-3, 'ul': 1e-3,
    'µL': 1e-3, 'μL': 1e-3, 'UL': 1e-3,
    'ml': 1.0, 'mL': 1.0, 'l': 1e3, 'L': 1e3,
}


def parse_conc_unit(unit_expr):
    """解析浓度单位表达式，返回换算系数 k，使 `数值 × k = 数值(μg/kg)`。

    支持乘除组合表达式，例如:
      'μg/kg'            -> 1
      'mg/kg'            -> 1e3
      'ng/g'             -> 1        (1 ng/g = 1 μg/kg)
      'μg/g'             -> 1e3
      '(μg/mL)×(μL)/g'   -> 1        (内标法常规单位，数值即 μg/kg)
      'mg/mL×μL/g'       -> 1e3
      'μg/mL'            -> 1e3      (质量/体积，按水密度 1 mL≈1 g 折算)
      'mg/L'             -> 1e3
    约定: 表达式最终量纲为 质量/质量（或 质量/体积，按水密度折算），
          无法识别时返回 None。
    """
    if not unit_expr or not isinstance(unit_expr, str):
        return None
    s = unit_expr.strip()
    # 归一化: 去空格、括号，乘号统一为 '*'
    s = s.replace(' ', '').replace('（', '').replace('）', '') \
         .replace('(', '').replace(')', '').replace('×', '*') \
         .replace('x', '*').replace('X', '*')
    if not s:
        return None
    conv = 1.0
    for tok in re.split(r'\*', s):
        sign = 1
        if tok.startswith('/'):
            sign = -1
            tok = tok[1:]
        if not tok:
            continue
        if '/' in tok:
            # 复合单位 mass/vol（如 μg/mL）或 mass/mass（如 g/μg）
            num, den = tok.split('/', 1)
            if num in _MASS_TO_G and den in _VOL_TO_ML:
                conv *= (_MASS_TO_G[num] ** sign) * (_VOL_TO_ML[den] ** (-sign))
            elif num in _MASS_TO_G and den in _MASS_TO_G:
                conv *= (_MASS_TO_G[num] ** sign) * (_MASS_TO_G[den] ** (-sign))
            elif num in _VOL_TO_ML and den in _MASS_TO_G:
                conv *= (_VOL_TO_ML[num] ** sign) * (_MASS_TO_G[den] ** (-sign))
            else:
                return None
        elif tok in _MASS_TO_G:
            conv *= _MASS_TO_G[tok] ** sign
        elif tok in _VOL_TO_ML:
            conv *= _VOL_TO_ML[tok] ** sign
        else:
            return None
    # conv 已将值折算为 克/毫升 组合量纲；统一折算到 μg/kg（质量比基准 ×1e9）
    return conv * 1e9



_SUP = {'⁰': '^0', '¹': '^1', '²': '^2', '³': '^3', '⁴': '^4',
        '⁵': '^5', '⁶': '^6', '⁷': '^7', '⁸': '^8', '⁹': '^9'}

def parse_threshold(s):
    """把阈值字符串解析为数值（取区间中值）；无法解析返回 None。
    支持: ≈240 / 7–65 / 50–200 / 1.0×10⁵ / 1–10 / — / —（母体≈1.0×10³） 等。"""
    if not isinstance(s, str):
        return None
    t = s.strip()
    if t in ('—', '-', '', '暂无', '无'):
        return None
    # 去掉括号注释与 ≈ / ~
    t = re.split(r'[（(]', t)[0]
    t = t.replace('≈', '').replace('~', '').replace(' ', '')
    if t in ('—', '-', ''):
        return None
    t = ''.join(_SUP.get(ch, ch) for ch in t)
    # 区间：用 en-dash 或 hyphen 分隔
    if '–' in t or '-' in t:
        vals = []
        for p in re.split(r'[–\-]', t):
            p = p.strip()
            if p:
                v = _to_num(p)
                if v is not None:
                    vals.append(v)
        return sum(vals) / len(vals) if vals else None
    return _to_num(t)

def _to_num(p):
    try:
        m = re.match(r'([0-9.]+)\s*[x×]\s*10\^?\s*([0-9]+)', p)
        if m:
            return float(m.group(1)) * (10 ** float(m.group(2)))
        return float(p)
    except Exception:
        return None

def compute_oav(rows, conc_unit=None):
    """对一批已富集行计算 OAV(气味活性值) 与 ROVA(相对气味活性值)，就地写入
    'oav' / 'roav' / 'oav_flag' / 'conc_ugkg'。

    浓度来源(优先级):
      1) 已算出的浓度 `conc`（内标法结果），按 `conc_unit`（每行的 conc_unit 或
         全局 conc_unit）换算到 μg/kg 后除以阈值 —— 这才是真实的 OAV;
      2) 回退用响应比 `rr`（无量纲）作为相对浓度，得到相对量纲 OAV（仅供相对比较）。
    公式:
      OAV  = C[μg/kg] / 阈值[μg/kg]          (OAV>1 即具气味活性)
      ROVA = OAV_i / max(OAV) × 100          (最高者=100；≥10 关键致香，≥1 潜在贡献)
    """
    pairs = []
    for r in rows:
        T = parse_threshold(r.get('thr'))
        if not T:
            continue
        conc = r.get('conc')
        rr = r.get('rr')
        val = None
        conc_ugkg = None
        if isinstance(conc, (int, float)) and conc is not None:
            unit = r.get('conc_unit') or conc_unit
            if unit:
                k = parse_conc_unit(unit)
                if k:
                    conc_ugkg = conc * k
                    val = conc_ugkg
            else:
                val = conc  # 未给单位时假定已为 μg/kg
                conc_ugkg = conc
        if val is None and isinstance(rr, (int, float)):
            val = rr  # 回退: 相对量纲（rr 无量纲）
        if val is None:
            continue
        pairs.append((r, val / T, conc_ugkg))
    if not pairs:
        return
    mx = max(v for _, v, _ in pairs)
    for r, v, cug in pairs:
        r['oav'] = round(v, 6)
        r['roav'] = round(v / mx * 100, 3) if mx else 0
        r['conc_ugkg'] = round(cug, 6) if isinstance(cug, (int, float)) else None
        if r['roav'] >= 10:
            r['oav_flag'] = '关键致香'
        elif r['roav'] >= 1:
            r['oav_flag'] = '潜在贡献'
        else:
            r['oav_flag'] = '—'


# ----------------------------------------------------------------------------
# 9. 内标法浓度计算
#    依据公式: C_物质 = (内标浓度 c₁ × 加入内标体积 V₁ × 响应比 rr) / 样品取样量 m
#    其中「响应比 rr」= 化合物峰面积 A / 内标峰面积 A₁（无量纲）。
#    因 rr 已包含内标峰面积 A₁，公式中不再单独除以 A₁，直接代入 rr 计算。
#    浓度单位由输入单位决定: (内标浓度单位 × 内标体积单位) / 样品取样量单位，
#    再经 parse_conc_unit 统一折算到 μg/kg。
# ----------------------------------------------------------------------------
def compute_concentration(c_is, v_is, response, sample_mass,
                          resp_istd=None, c_is_unit='μg/mL', v_is_unit='μL',
                          sample_mass_unit='g'):
    """单物质内标法浓度计算（直接使用响应比 rr 计算）。

    参数(任意可转 float 的数值或字符串):
      c_is           —— 内标浓度 c₁
      v_is           —— 加入内标体积 V₁
      response       —— 响应比 rr = 化合物峰面积 A / 内标峰面积 A₁（无量纲）
      sample_mass    —— 样品取样量 m
      c_is_unit      —— 内标浓度单位(如 μg/mL, mg/L, ng/μL)
      v_is_unit      —— 内标体积单位(如 μL, mL, L)
      sample_mass_unit —— 样品取样量单位(如 g, kg, mg)
    公式: C = (c₁ × V₁ × rr) ÷ m，结果折算到 μg/kg。
         响应比 rr 已包含内标峰面积 A₁，故公式中不再单独除以 A₁。
    返回: (conc_μgkg, unit_str) 或 (None, None)。
    """
    try:
        a = float(c_is); b = float(v_is); c = float(response); d = float(sample_mass)
    except (TypeError, ValueError):
        return None, None
    if d == 0 or not all(map(lambda x: x == x and abs(x) < float('inf'), (a, b, c, d))):
        return None, None
    raw = a * b * c / d
    # 单位换算: 把 (c_is_unit × v_is_unit) / sample_mass_unit 折算到 μg/kg
    unit_expr = '(%s)×(%s)/(%s)' % (c_is_unit, v_is_unit, sample_mass_unit)
    k = parse_conc_unit(unit_expr)
    if k is None:
        return None, unit_expr  # 单位无法识别，返回原组合单位供前端提示
    return raw * k, 'μg/kg'


def compute_concentration_list(c_is, v_is, sample_mass, items, **kw):
    """批量计算: items 为 [(name, response), ...]，返回 [(name, response, conc, unit), ...]。
    无法计算的项 conc 记为 None。"""
    out = []
    for name, resp in items:
        conc, unit = compute_concentration(c_is, v_is, resp, sample_mass, **kw)
        out.append((name, resp, conc, unit))
    return out


if __name__ == '__main__':
    # 自测
    tests = ["Butanoic acid", "2-Heptanone", "gamma-Nonalactone", "未知物 XYZ", "Limolene", "丁酸"]
    for t in tests:
        m = match_compound(t)
        print(f"{t:25s} -> {m['cn']:20s} {m['cat']:6s} 阈值={m['thr']:10s} [{m['match']}]")
    print('数据库规模:', len(COMPOUNDS), '检索索引:', len(DB))
    print('--- 阈值解析自测 ---')
    for s in ['≈240', '7–65', '50–200', '≈1.0×10⁵', '1–10', '≈0.07', '—', '—（母体≈1.0×10³）']:
        print(f'  {s!r:24s} -> {parse_threshold(s)}')


if __name__ == '__main__':
    # 自测
    tests = ["Butanoic acid", "2-Heptanone", "gamma-Nonalactone", "未知物 XYZ", "Limolene", "丁酸"]
    for t in tests:
        m = match_compound(t)
        print(f"{t:25s} -> {m['cn']:20s} {m['cat']:6s} 阈值={m['thr']:10s} [{m['match']}]")
    print('数据库规模:', len(COMPOUNDS), '检索索引:', len(DB))
